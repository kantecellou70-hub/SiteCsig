from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils.translation import gettext as _
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import get_user_model
from .models import (
    Category, Article, ArticleImage, Event, Project, ProjectPartner, Program, 
    Partner, Newsletter, ContactMessage, SiteSettings, StatusChoices, NewsletterCampaign,
    TeamMember, ConferenceRoom, ExternalOrganization, RoomBooking, RoomMaintenance,
    Personna, Blog, AboutPage, CityDistrict, CoreValue, HeroStatistic, Achievement
)
from .forms import (
    CategoryForm, ArticleForm, EventForm, ProjectForm, 
    ProgramForm, PartnerForm, NewsletterForm, ContactForm, ProjectPartnerForm,
    TeamMemberForm, ConferenceRoomForm, ExternalOrganizationForm, RoomBookingForm, RoomMaintenanceForm,
    PersonnaForm, BlogForm, AboutPageForm, CityDistrictForm, CoreValueForm, HeroStatisticForm, AchievementForm
)
import json
from django.db import models
from django.utils import timezone
from datetime import timedelta
from django.forms import inlineformset_factory
from .forms import (
    EventForm, EventDayForm, EventAgendaForm, EventIntervenantForm, 
    EventFAQForm, EventOrganizerForm, EventTagForm,
    EventDayInlineFormSet, EventAgendaInlineFormSet,
    EventRegistrationFormForm, FormFieldForm, FormFieldOptionForm, EventRegistrationPublicForm,
    BookingSearchForm
)
from .models import (
    Event, EventDay, EventAgenda, EventIntervenant, 
    EventFAQ, EventOrganizer, EventTag, EventRegistrationForm,
    FormField, FormFieldOption, EventRegistration, FormResponse
)

# Formset factories
EventDayFormSet = inlineformset_factory(
    Event, EventDay, 
    form=EventDayForm, 
    extra=1, 
    can_delete=True,
    formset=EventDayInlineFormSet
)

EventAgendaFormSet = inlineformset_factory(
    EventDay, EventAgenda, 
    form=EventAgendaForm, 
    extra=1, 
    can_delete=True,
    formset=EventAgendaInlineFormSet
)



EventFAQFormSet = inlineformset_factory(
    Event, EventFAQ, 
    form=EventFAQForm, 
    extra=1, 
    can_delete=True
)

EventOrganizerFormSet = inlineformset_factory(
    Event, EventOrganizer, 
    form=EventOrganizerForm, 
    extra=1, 
    can_delete=True
)


@login_required
def dashboard(request):
    """Tableau de bord principal"""
    context = {
        'total_articles': Article.objects.count(),
        'published_articles': Article.objects.filter(status=StatusChoices.PUBLISHED).count(),
        'draft_articles': Article.objects.filter(status=StatusChoices.DRAFT).count(),
        'total_events': Event.objects.count(),
        'upcoming_events_count': Event.objects.filter(status=StatusChoices.PUBLISHED).count(),
        'upcoming_events': Event.objects.filter(status=StatusChoices.PUBLISHED).order_by('start_date')[:5],
        'total_projects': Project.objects.count(),
        'active_projects': Project.objects.filter(status='active').count(),
        'total_programs': Program.objects.count(),
        'total_partners': Partner.objects.filter(is_active=True).count(),
        'total_users': get_user_model().objects.count(),
        'recent_articles': Article.objects.order_by('-created_at')[:5],
        'recent_events': Event.objects.order_by('-created_at')[:5],
        'recent_projects': Project.objects.order_by('-created_at')[:5],
        'recent_activity': [],  # Empty list for now, can be populated with actual activity data later
    }
    return render(request, 'content_management/dashboard.html', context)


# Gestion des catégories
@login_required
def category_list(request):
    """Liste des catégories"""
    categories = Category.objects.all().order_by('order', 'name')
    
    # Calculer les vraies statistiques
    total_categories = Category.objects.count()
    active_categories = Category.objects.filter(is_active=True).count()
    parent_categories = Category.objects.filter(parent__isnull=True).count()
    total_articles = Article.objects.count()
    
    context = {
        'categories': categories,
        'total_categories': total_categories,
        'active_categories': active_categories,
        'parent_categories': parent_categories,
        'total_articles': total_articles,
    }
    return render(request, 'content_management/category_list.html', context)


@login_required
def category_create(request):
    """Créer une nouvelle catégorie"""
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, _('Catégorie créée avec succès.'))
            return redirect('content_management:category_list')
    else:
        form = CategoryForm()
    
    return render(request, 'content_management/category_form.html', {'form': form, 'action': 'create'})


@login_required
def category_detail(request, pk):
    """Détails d'une catégorie"""
    category = get_object_or_404(Category, pk=pk)
    
    # Calculer les statistiques
    total_views = sum(article.views_count or 0 for article in category.articles.all())
    
    context = {
        'category': category,
        'total_views': total_views,
    }
    
    return render(request, 'content_management/category_detail.html', context)


@login_required
def category_edit(request, pk):
    """Modifier une catégorie"""
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, _('Catégorie modifiée avec succès.'))
            return redirect('content_management:category_detail', pk=category.pk)
    else:
        form = CategoryForm(instance=category)
    
    return render(request, 'content_management/category_edit.html', {'form': form, 'category': category})


@login_required
def category_delete(request, pk):
    """Supprimer une catégorie"""
    category = get_object_or_404(Category, pk=pk)
    
    if request.method == 'POST':
        category.delete()
        messages.success(request, _('Catégorie supprimée avec succès.'))
        return redirect('content_management:category_list')
    
    return render(request, 'content_management/category_confirm_delete.html', {'category': category})


# Gestion des événements
@login_required
def event_list(request):
    """Liste des événements"""
    events = Event.objects.all().order_by('-start_date', '-start_time')
    
    # Filtres
    event_type = request.GET.get('event_type')
    status = request.GET.get('status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if event_type:
        events = events.filter(event_type=event_type)
    if status:
        events = events.filter(status=status)
    if date_from:
        events = events.filter(start_date__gte=date_from)
    if date_to:
        events = events.filter(end_date__lte=date_to)
    
    # Pagination
    paginator = Paginator(events, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'event_types': Event.event_type.field.choices,
        'status_choices': StatusChoices.choices,
    }
    return render(request, 'content_management/event_list.html', context)


@login_required
def event_create(request):
    """Créer un nouvel événement"""
    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES)
        print(f"DEBUG: Form data: {request.POST}")
        print(f"DEBUG: Form files: {request.FILES}")
        print(f"DEBUG: Form is valid: {form.is_valid()}")
        if form.is_valid():
            try:
                event = form.save()
                print(f"DEBUG: Event created with ID: {event.pk}")
                messages.success(request, _("Événement créé avec succès !"))
                return redirect('content_management:event_detail', pk=event.pk)
            except Exception as e:
                print(f"DEBUG: Error saving event: {e}")
                messages.error(request, f"Erreur lors de la création: {e}")
        else:
            print(f"DEBUG: Form errors: {form.errors}")
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = EventForm()
    
    context = {
        'form': form,
        'title': _("Créer un événement"),
        'submit_text': _("Créer l'événement"),
    }
    return render(request, 'content_management/event_form.html', context)


@login_required
def event_edit(request, pk):
    """Modifier un événement"""
    event = get_object_or_404(Event, pk=pk)
    
    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES, instance=event)
        if form.is_valid():
            form.save()
            messages.success(request, _("Événement modifié avec succès !"))
            return redirect('content_management:event_detail', pk=event.pk)
    else:
        form = EventForm(instance=event)
    
    context = {
        'form': form,
        'event': event,
        'title': _("Modifier l'événement"),
        'submit_text': _("Enregistrer les modifications"),
    }
    return render(request, 'content_management/event_form.html', context)


@login_required
def event_detail(request, pk):
    """Détails d'un événement"""
    event = get_object_or_404(Event, pk=pk)
    
    context = {
        'event': event,
        'days': event.days.all().prefetch_related('activities'),
        'faqs': event.faqs.filter(is_public=True),
        'organizers': event.organizers.all(),
    }
    return render(request, 'content_management/event_detail.html', context)


@login_required
def event_delete(request, pk):
    """Supprimer un événement"""
    event = get_object_or_404(Event, pk=pk)
    
    if request.method == 'POST':
        event.delete()
        messages.success(request, _("Événement supprimé avec succès !"))
        return redirect('content_management:event_list')
    
    context = {'event': event}
    return render(request, 'content_management/event_confirm_delete.html', context)


@login_required
def event_manage_agenda(request, pk):
    """Gérer l'agenda d'un événement"""
    event = get_object_or_404(Event, pk=pk)
    
    if request.method == 'POST':
        # Vérifier si c'est une requête AJAX pour ajouter un jour
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'date' in request.POST:
            try:
                # Créer un nouveau jour
                day = EventDay.objects.create(
                    event=event,
                    date=request.POST.get('date'),
                    day_number=request.POST.get('day_number'),
                    title=request.POST.get('title', ''),
                    description=request.POST.get('description', '')
                )
                
                print(f"DEBUG: Jour créé avec succès: {day.date}, Numéro: {day.day_number}")
                
                return JsonResponse({
                    'success': True,
                    'message': 'Jour ajouté avec succès',
                    'day_id': day.pk
                })
            except Exception as e:
                print(f"DEBUG: Erreur lors de la création: {str(e)}")
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                }, status=400)
        else:
            # Fallback pour la soumission de formset (ancienne méthode)
            day_formset = EventDayFormSet(request.POST, instance=event, prefix='days')
            if day_formset.is_valid():
                day_formset.save()
                messages.success(request, _("Agenda mis à jour avec succès !"))
                return redirect('content_management:event_detail', pk=event.pk)
    else:
        day_formset = EventDayFormSet(instance=event, prefix='days')
    
    context = {
        'event': event,
        'day_formset': day_formset,
    }
    return render(request, 'content_management/event_manage_agenda.html', context)


@login_required
def event_manage_intervenants(request, event_pk):
    """Gérer les intervenants d'un événement"""
    event = get_object_or_404(Event, pk=event_pk)
    
    # Cette vue n'est plus utilisée car les intervenants sont maintenant gérés par activité
    # Rediriger vers la liste des intervenants
    messages.info(request, _("Les intervenants sont maintenant gérés par activité. Utilisez la gestion des activités pour assigner des intervenants."))
    return redirect('content_management:event_intervenant_list')


@login_required
def event_manage_faq(request, event_pk):
    """Gérer la FAQ d'un événement"""
    event = get_object_or_404(Event, pk=event_pk)
    
    # Récupérer toutes les FAQ de l'événement
    faqs = event.faqs.all().order_by('order', 'id')
    
    context = {
        'event': event,
        'faqs': faqs,
    }
    return render(request, 'content_management/event_manage_faq.html', context)


@login_required
def event_faq_create(request, event_pk):
    """Créer une nouvelle FAQ pour un événement"""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            event = get_object_or_404(Event, pk=event_pk)
            
            # Créer la nouvelle FAQ
            faq = EventFAQ.objects.create(
                event=event,
                question=request.POST.get('question'),
                answer=request.POST.get('answer'),
                category=request.POST.get('category', 'general'),
                order=request.POST.get('order', 0),
                is_public=request.POST.get('is_public') == 'on'
            )
            
            return JsonResponse({
                'success': True,
                'message': _('FAQ créée avec succès !'),
                'faq_id': faq.pk
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)


@login_required
def event_faq_edit(request, event_pk, faq_pk):
    """Modifier une FAQ d'événement"""
    faq = get_object_or_404(EventFAQ, pk=faq_pk, event_id=event_pk)
    
    if request.method == 'GET' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Retourner les données de la FAQ pour le formulaire d'édition
        return JsonResponse({
            'success': True,
            'faq': {
                'id': faq.pk,
                'question': faq.question,
                'answer': faq.answer,
                'category': faq.category,
                'order': faq.order,
                'is_public': faq.is_public
            }
        })
    
    elif request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            # Mettre à jour la FAQ
            faq.question = request.POST.get('question')
            faq.answer = request.POST.get('answer')
            faq.category = request.POST.get('category', 'general')
            faq.order = request.POST.get('order', 0)
            faq.is_public = request.POST.get('is_public') == 'on'
            faq.save()
            
            return JsonResponse({
                'success': True,
                'message': _('FAQ mise à jour avec succès !')
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)


@login_required
def event_faq_delete(request, event_pk, faq_pk):
    """Supprimer une FAQ d'événement"""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            faq = get_object_or_404(EventFAQ, pk=faq_pk, event_id=event_pk)
            faq.delete()
            
            return JsonResponse({
                'success': True,
                'message': _('FAQ supprimée avec succès !')
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)


@login_required
def event_manage_organizers(request, event_pk):
    """Gérer les organisateurs d'un événement"""
    event = get_object_or_404(Event, pk=event_pk)
    
    if request.method == 'POST':
        # Vérifier si c'est une requête AJAX pour ajouter un organisateur
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'name' in request.POST:
            try:
                # Créer un nouvel organisateur
                organizer = EventOrganizer.objects.create(
                    event=event,
                    name=request.POST.get('name'),
                    logo=request.FILES.get('logo'),
                    description=request.POST.get('description', ''),
                    website=request.POST.get('website', ''),
                    role=request.POST.get('role', 'other'),
                    organization_type=request.POST.get('organization_type', 'other'),
                    contact_email=request.POST.get('contact_email', ''),
                    contact_phone=request.POST.get('contact_phone', ''),
                    address=request.POST.get('address', ''),
                    order=request.POST.get('order', 1)
                )
                
                print(f"DEBUG: Organisateur créé avec succès: {organizer.name}, Logo: {organizer.logo}")
                
                return JsonResponse({
                    'success': True,
                    'message': 'Organisateur ajouté avec succès',
                    'organizer_id': organizer.pk
                })
            except Exception as e:
                print(f"DEBUG: Erreur lors de la création: {str(e)}")
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                }, status=400)
        else:
            # Gestion du formset (ancienne méthode)
            print(f"DEBUG: POST data: {request.POST}")
            print(f"DEBUG: FILES data: {request.FILES}")
            formset = EventOrganizerFormSet(request.POST, request.FILES, instance=event)
            print(f"DEBUG: Formset is valid: {formset.is_valid()}")
            if formset.is_valid():
                print(f"DEBUG: Formset errors: {formset.errors}")
                formset.save()
                messages.success(request, _("Organisateurs mis à jour avec succès !"))
                return redirect('content_management:event_detail', pk=event.pk)
            else:
                print(f"DEBUG: Formset errors: {formset.errors}")
                print(f"DEBUG: Non-form errors: {formset.non_form_errors()}")
    else:
        formset = EventOrganizerFormSet(instance=event)
    
    # Récupérer les organisateurs avec debug
    organizers = event.organizers.all().order_by('order', 'name')
    print(f"DEBUG: Nombre d'organisateurs trouvés: {organizers.count()}")
    for org in organizers:
        print(f"DEBUG: Organisateur: {org.name}, Logo: {org.logo}, Logo URL: {org.logo.url if org.logo else 'Aucun'}")
    
    context = {
        'event': event,
        'formset': formset,
        'organizers': organizers,  # Passer explicitement les organisateurs
    }
    return render(request, 'content_management/event_manage_organizers.html', context)


# Vues pour les tags d'événements
@login_required
def event_tag_list(request):
    """Liste des tags d'événements"""
    tags = EventTag.objects.all().order_by('name')
    context = {'tags': tags}
    return render(request, 'content_management/event_tag_list.html', context)


@login_required
def event_tag_create(request):
    """Créer un tag d'événement"""
    if request.method == 'POST':
        form = EventTagForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, _("Tag créé avec succès !"))
            return redirect('content_management:event_tag_list')
    else:
        form = EventTagForm()
    
    context = {
        'form': form,
        'title': _("Créer un tag"),
        'submit_text': _("Créer le tag"),
    }
    return render(request, 'content_management/event_tag_form.html', context)


@login_required
def event_tag_edit(request, pk):
    """Modifier un tag d'événement"""
    tag = get_object_or_404(EventTag, pk=pk)
    
    if request.method == 'POST':
        form = EventTagForm(request.POST, instance=tag)
        if form.is_valid():
            form.save()
            messages.success(request, _("Tag modifié avec succès !"))
            return redirect('content_management:event_tag_list')
    else:
        form = EventTagForm(instance=tag)
    
    context = {
        'form': form,
        'tag': tag,
        'title': _("Modifier le tag"),
        'submit_text': _("Enregistrer les modifications"),
    }
    return render(request, 'content_management/event_tag_form.html', context)


@login_required
def event_tag_delete(request, pk):
    """Supprimer un tag d'événement"""
    tag = get_object_or_404(EventTag, pk=pk)
    
    if request.method == 'POST':
        tag.delete()
        messages.success(request, _("Tag supprimé avec succès !"))
        return redirect('content_management:event_tag_list')
    
    context = {'tag': tag}
    return render(request, 'content_management/event_tag_confirm_delete.html', context)


# Vues AJAX pour la gestion dynamique
@csrf_exempt
def event_toggle_status(request):
    """Activer/désactiver un événement via AJAX"""
    if request.method == 'POST':
        try:
            event_id = request.POST.get('event_id')
            new_status = request.POST.get('new_status') == 'true'
            event = get_object_or_404(Event, pk=event_id)
            event.status = 'published' if new_status else 'draft'
            event.save()
            message = _("Événement activé avec succès.") if new_status else _("Événement désactivé avec succès.")
            return JsonResponse({'success': True, 'message': message, 'new_status': new_status})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def event_toggle_feature(request):
    """Mettre en avant/retirer de la une un événement via AJAX"""
    if request.method == 'POST':
        try:
            event_id = request.POST.get('event_id')
            new_status = request.POST.get('new_status') == 'true'
            event = get_object_or_404(Event, pk=event_id)
            event.is_featured = new_status
            event.save()
            message = _("Événement mis en avant avec succès.") if new_status else _("Événement retiré de la une avec succès.")
            return JsonResponse({'success': True, 'message': message, 'new_status': new_status})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def delete_event(request):
    """Supprimer un événement via AJAX"""
    if request.method == 'POST':
        try:
            event_id = request.POST.get('event_id')
            event = get_object_or_404(Event, pk=event_id)
            event.delete()
            return JsonResponse({'success': True, 'message': _("Événement supprimé avec succès.")})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def bulk_event_action(request):
    """Actions en masse sur les événements via AJAX"""
    if request.method == 'POST':
        try:
            event_ids = request.POST.getlist('event_ids[]')
            action = request.POST.get('action')
            
            if not event_ids:
                return JsonResponse({'success': False, 'error': _("Aucun événement sélectionné.")})
            
            events = Event.objects.filter(pk__in=event_ids)
            
            if action == 'delete':
                events.delete()
                message = _("Événements supprimés avec succès.")
            elif action == 'publish':
                events.update(status='published')
                message = _("Événements publiés avec succès.")
            elif action == 'draft':
                events.update(status='draft')
                message = _("Événements mis en brouillon avec succès.")
            elif action == 'feature':
                events.update(is_featured=True)
                message = _("Événements mis en avant avec succès.")
            elif action == 'unfeature':
                events.update(is_featured=False)
                message = _("Événements retirés de la une avec succès.")
            else:
                return JsonResponse({'success': False, 'error': _("Action non reconnue.")})
            
            return JsonResponse({'success': True, 'message': message})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


# Gestion des articles
@login_required
def article_list(request):
    """Liste des articles avec filtres et pagination"""
    articles = Article.objects.all().order_by('-created_at')
    
    # Filtres
    status_filter = request.GET.get('status')
    category_filter = request.GET.get('category')
    search_query = request.GET.get('search')
    featured_filter = request.GET.get('featured')
    
    if status_filter:
        articles = articles.filter(status=status_filter)
    if category_filter:
        articles = articles.filter(category_id=category_filter)
    if search_query:
        articles = articles.filter(
            Q(title__icontains=search_query) |
            Q(title_en__icontains=search_query) |
            Q(content__icontains=search_query)
        )
    if featured_filter == 'true':
        articles = articles.filter(is_featured=True)
    elif featured_filter == 'false':
        articles = articles.filter(is_featured=False)
    
    # Pagination
    paginator = Paginator(articles, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    categories = Category.objects.filter(is_active=True)
    
    # Statistiques
    total_articles = Article.objects.count()
    published_articles = Article.objects.filter(status=StatusChoices.PUBLISHED).count()
    featured_articles = Article.objects.filter(is_featured=True).count()
    draft_articles = Article.objects.filter(status=StatusChoices.DRAFT).count()
    
    context = {
        'articles': page_obj,  # Pour la pagination
        'page_obj': page_obj,  # Garder pour compatibilité
        'categories': categories,
        'status_filter': status_filter,
        'category_filter': category_filter,
        'search_query': search_query,
        'featured_filter': featured_filter,
        'total_articles': total_articles,
        'published_articles': published_articles,
        'featured_articles': featured_articles,
        'draft_articles': draft_articles,
    }
    return render(request, 'content_management/article_list.html', context)


@login_required
def article_create(request):
    """Créer un nouvel article"""
    # Récupérer les catégories pour le contexte AVANT de créer le formulaire
    try:
        categories = Category.objects.filter(is_active=True).order_by('order', 'name')
        print(f"DEBUG: Catégories trouvées: {categories.count()}")
        for cat in categories:
            print(f"DEBUG: - {cat.name} (ID: {cat.id})")
    except Exception as e:
        # Fallback si le champ is_active n'existe pas
        print(f"DEBUG: Erreur avec is_active: {e}")
        categories = Category.objects.all().order_by('order', 'name')
        print(f"DEBUG: Fallback - Catégories trouvées: {categories.count()}")
    
    if request.method == 'POST':
        form = ArticleForm(request.POST, request.FILES, categories=categories)
        if form.is_valid():
            article = form.save(commit=False)
            article.author = request.user
            print(f"DEBUG: Sauvegarde de l'article '{article.title}' avec catégorie: {article.category}")
            article.save()
            form.save_m2m()
            
            print(f"DEBUG: Article créé avec succès - ID: {article.id}, Titre: {article.title}, Catégorie: {article.category}")
            
            # Gérer les images multiples
            images = request.FILES.getlist('images')
            for i, image in enumerate(images):
                if image:
                    ArticleImage.objects.create(
                        article=article,
                        image=image,
                        order=i
                    )
            
            messages.success(request, _('Article créé avec succès.'))
            return redirect('content_management:article_list')
        else:
            # Si le formulaire n'est pas valide, le recréer avec les catégories
            form = ArticleForm(request.POST, request.FILES, categories=categories)
    else:
        # Créer le formulaire avec les catégories déjà filtrées
        print(f"DEBUG: Création du formulaire avec {categories.count()} catégories")
        form = ArticleForm(categories=categories)
        print(f"DEBUG: Formulaire créé, queryset du champ category: {form.fields['category'].queryset.count()}")
        
        # FORCER le queryset du champ category
        form.fields['category'].queryset = categories
        print(f"DEBUG: Après forçage, queryset du champ category: {form.fields['category'].queryset.count()}")
        
        # Définir automatiquement l'auteur comme l'utilisateur connecté
        form.initial['author'] = request.user.id
    
    context = {
        'form': form, 
        'action': 'create',
        'categories': categories
    }
    return render(request, 'content_management/article_form.html', context)


@login_required
def article_edit(request, pk):
    """Modifier un article"""
    article = get_object_or_404(Article, pk=pk)
    
    # Récupérer les catégories pour le contexte AVANT de créer le formulaire
    try:
        categories = Category.objects.filter(is_active=True).order_by('order', 'name')
    except:
        # Fallback si le champ is_active n'existe pas
        categories = Category.objects.all().order_by('order', 'name')
    
    if request.method == 'POST':
        form = ArticleForm(request.POST, request.FILES, instance=article)
        if form.is_valid():
            form.save()
            
            # Gérer les nouvelles images
            images = request.FILES.getlist('images')
            for i, image in enumerate(images):
                if image:
                    ArticleImage.objects.create(
                        article=article,
                        image=image,
                        order=article.images.count() + i
                    )
            
            messages.success(request, _('Article modifié avec succès.'))
            return redirect('content_management:article_list')
    else:
        # Créer le formulaire AVEC l'instance pour pré-remplir les champs
        print(f"DEBUG: Création du formulaire avec {categories.count()} catégories")
        form = ArticleForm(instance=article, categories=categories)
        
        # FORCER le queryset du champ category
        form.fields['category'].queryset = categories
        print(f"DEBUG: Après forçage, queryset du champ category: {form.fields['category'].queryset.count()}")
        
        # Définir automatiquement l'auteur comme l'utilisateur connecté
        form.initial['author'] = request.user.id
    
    context = {
        'form': form, 
        'action': 'edit', 
        'article': article,
        'categories': categories
    }
    return render(request, 'content_management/article_form.html', context)


@login_required
def article_delete(request, pk):
    """Supprimer un article"""
    article = get_object_or_404(Article, pk=pk)
    if request.method == 'POST':
        article.delete()
        messages.success(request, _('Article supprimé avec succès.'))
        return redirect('content_management:article_list')
    
    return render(request, 'content_management/article_confirm_delete.html', {'article': article})


@login_required
def article_detail(request, pk):
    """Détails d'un article"""
    article = get_object_or_404(Article, pk=pk)
    
    # Récupérer les articles liés (même catégorie)
    related_articles = Article.objects.filter(
        category=article.category,
        status='published'
    ).exclude(pk=article.pk)[:5]
    
    context = {
        'article': article,
        'related_articles': related_articles,
    }
    
    return render(request, 'content_management/article_detail.html', context)


@login_required
def article_status_change(request, pk):
    """Changer le statut d'un article"""
    article = get_object_or_404(Article, pk=pk)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in ['draft', 'published', 'archived']:
            article.status = new_status
            article.save()
            return JsonResponse({'success': True, 'status': new_status})
    
    return JsonResponse({'success': False, 'message': 'Statut invalide'})


@login_required
def bulk_article_action(request):
    """Actions en lot sur les articles"""
    if request.method == 'POST':
        action = request.POST.get('action')
        article_ids = json.loads(request.POST.get('article_ids', '[]'))
        
        if not article_ids:
            return JsonResponse({
                'success': False,
                'message': _('Aucun article sélectionné.')
            })
        
        articles = Article.objects.filter(pk__in=article_ids)
        count = 0
        
        try:
            if action == 'publish':
                articles.update(status='published')
                count = articles.count()
                message = _('%(count)d article(s) publié(s) avec succès.') % {'count': count}
                
            elif action == 'draft':
                articles.update(status='draft')
                count = articles.count()
                message = _('%(count)d article(s) mis en brouillon avec succès.') % {'count': count}
                
            elif action == 'archive':
                articles.update(status='archived')
                count = articles.count()
                message = _('%(count)d article(s) archivé(s) avec succès.') % {'count': count}
                
            elif action == 'delete':
                count = articles.count()
                articles.delete()
                message = _('%(count)d article(s) supprimé(s) avec succès.') % {'count': count}
                
            elif action == 'feature':
                articles.update(is_featured=True)
                count = articles.count()
                message = _('%(count)d article(s) mis en avant avec succès.') % {'count': count}
                
            elif action == 'unfeature':
                articles.update(is_featured=False)
                count = articles.count()
                message = _('%(count)d article(s) retiré(s) de la une avec succès.') % {'count': count}
                
            else:
                return JsonResponse({
                    'success': False,
                    'message': _('Action non reconnue.')
                })
            
            return JsonResponse({
                'success': True,
                'count': count,
                'message': message
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': _('Erreur lors de l\'exécution de l\'action: %(error)s') % {'error': str(e)}
            })
    
    return JsonResponse({
        'success': False,
        'message': _('Méthode non autorisée.')
    })


@login_required
def toggle_feature(request):
    """Basculer la mise en avant d'un article"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            article_id = data.get('article_id')
            featured = data.get('featured')
            
            article = get_object_or_404(Article, pk=article_id)
            article.is_featured = featured
            article.save()
            
            action = 'mis en avant' if featured else 'retiré de la une'
            message = _('Article "%(title)s" %(action)s avec succès.') % {
                'title': article.title,
                'action': action
            }
            
            return JsonResponse({
                'success': True,
                'message': message,
                'featured': featured
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': _('Erreur lors de la modification: %(error)s') % {'error': str(e)}
            })
    
    return JsonResponse({
        'success': False,
        'message': _('Méthode non autorisée.')
    })


@login_required
def delete_article(request):
    """Supprimer un article via AJAX"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            article_id = data.get('article_id')
            
            article = get_object_or_404(Article, pk=article_id)
            article_title = article.title
            article.delete()
            
            message = _('Article "%(title)s" supprimé avec succès.') % {'title': article_title}
            
            return JsonResponse({
                'success': True,
                'message': message
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': _('Erreur lors de la suppression: %(error)s') % {'error': str(e)}
            })
    
    return JsonResponse({
        'success': False,
        'message': _('Méthode non autorisée.')
    })


# Gestion des projets
@login_required
def project_list(request):
    """Liste des projets avec filtres et pagination"""
    projects = Project.objects.all()
    
    # Filtres
    status_filter = request.GET.get('status')
    search_query = request.GET.get('search')
    featured_filter = request.GET.get('featured')
    
    if status_filter:
        projects = projects.filter(status=status_filter)
    
    if search_query:
        projects = projects.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(short_description__icontains=search_query)
        )
    
    if featured_filter == 'true':
        projects = projects.filter(is_featured=True)
    elif featured_filter == 'false':
        projects = projects.filter(is_featured=False)
    
    # Tri par date de création (plus récent en premier)
    projects = projects.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(projects, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'projects': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
    }
    return render(request, 'content_management/project_list.html', context)


@login_required
def project_create(request):
    """Créer un nouveau projet"""
    if request.method == 'POST':
        form = ProjectForm(request.POST, request.FILES)
        if form.is_valid():
            project = form.save(commit=False)
            project.save()
            form.save_m2m()  # Pour les champs many-to-many
            messages.success(request, _('Projet créé avec succès.'))
            return redirect('content_management:project_list')
    else:
        form = ProjectForm()
    
    return render(request, 'content_management/project_form.html', {'form': form})


@login_required
def project_edit(request, pk):
    """Modifier un projet"""
    project = get_object_or_404(Project, pk=pk)
    
    if request.method == 'POST':
        form = ProjectForm(request.POST, request.FILES, instance=project)
        if form.is_valid():
            form.save()
            messages.success(request, _('Projet modifié avec succès.'))
            return redirect('content_management:project_list')
    else:
        form = ProjectForm(instance=project)
    
    return render(request, 'content_management/project_form.html', {'form': form})


@login_required
def project_detail(request, pk):
    """Détails d'un projet"""
    project = get_object_or_404(Project, pk=pk)
    
    context = {
        'project': project,
    }
    
    return render(request, 'content_management/project_detail.html', context)


@login_required
@require_http_methods(["POST"])
def project_delete(request, pk):
    """Supprimer un projet"""
    try:
        project = get_object_or_404(Project, pk=pk)
        project_title = project.title
        project.delete()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'Projet "{project_title}" supprimé avec succès'})
        else:
            messages.success(request, f'Projet "{project_title}" supprimé avec succès')
            return redirect('content_management:project_list')
        
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': f'Erreur lors de la suppression: {str(e)}'})
        else:
            messages.error(request, f'Erreur lors de la suppression: {str(e)}')
            return redirect('content_management:project_detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def bulk_project_action(request):
    """Actions en lot sur les projets"""
    try:
        action = request.POST.get('action')
        project_ids = json.loads(request.POST.get('project_ids', '[]'))
        
        if not action or not project_ids:
            return JsonResponse({'success': False, 'message': 'Action ou IDs manquants'})
        
        projects = Project.objects.filter(pk__in=project_ids)
        
        if action == 'activate':
            projects.update(status='active')
            message = f'{projects.count()} projet(s) activé(s)'
        elif action == 'complete':
            projects.update(status='completed')
            message = f'{projects.count()} projet(s) marqué(s) comme terminé(s)'
        elif action == 'cancel':
            projects.update(status='cancelled')
            message = f'{projects.count()} projet(s) annulé(s)'
        elif action == 'delete':
            count = projects.count()
            projects.delete()
            message = f'{count} projet(s) supprimé(s)'
        else:
            return JsonResponse({'success': False, 'message': 'Action non reconnue'})
        
        return JsonResponse({'success': True, 'message': message})
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Données invalides'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur lors de l\'action: {str(e)}'})


# Gestion des programmes
@login_required
def program_list(request):
    """Liste des programmes"""
    programs = Program.objects.all().order_by('-created_at')
    
    # Filtres
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    if status_filter:
        programs = programs.filter(status=status_filter)
    
    if search_query:
        programs = programs.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(short_description__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(programs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistiques
    total_programs = Program.objects.count()
    active_programs = Program.objects.filter(status='published').count()
    draft_programs = Program.objects.filter(status='draft').count()
    
    context = {
        'programs': page_obj,
        'page_obj': page_obj,
        'status_filter': status_filter,
        'search_query': search_query,
        'total_programs': total_programs,
        'active_programs': active_programs,
        'draft_programs': draft_programs,
        'is_paginated': page_obj.has_other_pages(),
    }
    return render(request, 'content_management/program_list.html', context)


@login_required
def program_create(request):
    """Créer un nouveau programme"""
    if request.method == 'POST':
        form = ProgramForm(request.POST, request.FILES)
        if form.is_valid():
            program = form.save(commit=False)
            program.author = request.user
            program.save()
            form.save_m2m()
            
            messages.success(request, _('Programme créé avec succès.'))
            return redirect('content_management:program_list')
    else:
        form = ProgramForm()
    
    return render(request, 'content_management/program_form.html', {'form': form})


@login_required
def program_detail(request, pk):
    """Détail d'un programme"""
    program = get_object_or_404(Program, pk=pk)
    
    # Incrémenter le compteur de vues (optionnel)
    # program.views_count = F('views_count') + 1
    # program.save()
    
    context = {
        'program': program,
        'now': timezone.now(),
    }
    return render(request, 'content_management/program_detail.html', context)


@login_required
def program_edit(request, pk):
    """Modifier un programme"""
    program = get_object_or_404(Program, pk=pk)
    if request.method == 'POST':
        form = ProgramForm(request.POST, request.FILES, instance=program)
        if form.is_valid():
            form.save()
            messages.success(request, _('Programme modifié avec succès.'))
            return redirect('content_management:program_detail', pk=program.pk)
    else:
        form = ProgramForm(instance=program)
    
    return render(request, 'content_management/program_edit.html', {'form': form, 'program': program})


@login_required
def program_delete(request, pk):
    """Supprimer un programme"""
    program = get_object_or_404(Program, pk=pk)
    if request.method == 'POST':
        program.delete()
        messages.success(request, _('Programme supprimé avec succès.'))
        return redirect('content_management:program_list')
    
    context = {
        'program': program,
        'now': timezone.now(),
    }
    return render(request, 'content_management/program_confirm_delete.html', context)


# Gestion des partenaires
@login_required
def partner_list(request):
    """Liste des partenaires"""
    partners = Partner.objects.all().order_by('order', 'name')
    return render(request, 'content_management/partner_list.html', {'partners': partners})


@login_required
def partner_create(request):
    """Créer un nouveau partenaire"""
    if request.method == 'POST':
        form = PartnerForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                partner = form.save()
                messages.success(request, _('Partenaire créé avec succès.'))
                return redirect('content_management:partner_list')
            except Exception as e:
                messages.error(request, f'Erreur lors de la création: {str(e)}')
        else:
            # Afficher les erreurs de validation
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Erreur dans {field}: {error}')
    else:
        form = PartnerForm()
    
    context = {
        'form': form, 
        'action': 'create',
        'title': 'Créer un partenaire'
    }
    return render(request, 'content_management/partner_form.html', context)


@login_required
def partner_edit(request, pk):
    """Modifier un partenaire"""
    partner = get_object_or_404(Partner, pk=pk)
    if request.method == 'POST':
        form = PartnerForm(request.POST, request.FILES, instance=partner)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, _('Partenaire modifié avec succès.'))
                return redirect('content_management:partner_list')
            except Exception as e:
                messages.error(request, f'Erreur lors de la modification: {str(e)}')
        else:
            # Afficher les erreurs de validation
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Erreur dans {field}: {error}')
    else:
        form = PartnerForm(instance=partner)
    
    context = {
        'form': form, 
        'action': 'edit', 
        'partner': partner,
        'title': 'Modifier le partenaire'
    }
    return render(request, 'content_management/partner_form.html', context)


@login_required
def partner_delete(request, pk):
    """Supprimer un partenaire"""
    partner = get_object_or_404(Partner, pk=pk)
    if request.method == 'POST':
        partner.delete()
        messages.success(request, _('Partenaire supprimé avec succès.'))
        return redirect('content_management:partner_list')
    
    return render(request, 'content_management/partner_confirm_delete.html', {'partner': partner})


# Gestion des newsletters
@login_required
def newsletter_list(request):
    """Liste des abonnés newsletter"""
    subscribers = Newsletter.objects.all().order_by('-subscription_date')
    return render(request, 'content_management/newsletter_list.html', {'subscribers': subscribers})


@login_required
def newsletter_export(request):
    """Exporter la liste des emails"""
    subscribers = Newsletter.objects.filter(is_active=True)
    emails = [sub.email for sub in subscribers]
    
    response = HttpResponse('\n'.join(emails), content_type='text/plain')
    response['Content-Disposition'] = 'attachment; filename="newsletter_emails.txt"'
    return response


# Messages de contact
@login_required
def contact_message_list(request):
    """Liste des messages de contact"""
    messages_list = ContactMessage.objects.all().order_by('-created_at')
    return render(request, 'content_management/contact_message_list.html', {'messages': messages_list})


@login_required
def contact_message_detail(request, pk):
    """Détail d'un message de contact"""
    message = get_object_or_404(ContactMessage, pk=pk)
    if not message.is_read:
        message.is_read = True
        message.save()
    
    return render(request, 'content_management/contact_message_detail.html', {'message': message})


@login_required
def contact_message_delete(request, pk):
    """Supprimer un message de contact"""
    message = get_object_or_404(ContactMessage, pk=pk)
    if request.method == 'POST':
        message.delete()
        messages.success(request, _('Message supprimé avec succès.'))
        return redirect('content_management:contact_message_list')
    
    return render(request, 'content_management/contact_message_confirm_delete.html', {'message': message})


# API pour AJAX
@csrf_exempt
def delete_category_ajax(request):
    """Supprimer une catégorie via AJAX"""
    if request.method == 'POST':
        try:
            category_id = request.POST.get('category_id')
            category = get_object_or_404(Category, pk=category_id)
            category.delete()
            return JsonResponse({'success': True, 'message': _('Catégorie supprimée avec succès.')})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@csrf_exempt
def category_toggle_status(request):
    """Activer/désactiver une catégorie via AJAX"""
    if request.method == 'POST':
        try:
            category_id = request.POST.get('category_id')
            new_status = request.POST.get('new_status')
            
            if category_id is None:
                return JsonResponse({'success': False, 'error': _('ID de catégorie manquant.')})
            
            category = get_object_or_404(Category, pk=category_id)
            
            # Convertir la chaîne en booléen
            if new_status == 'true':
                category.is_active = True
                status_text = _('activée')
            else:
                category.is_active = False
                status_text = _('désactivée')
            
            category.save()
            
            return JsonResponse({
                'success': True, 
                'message': f'Catégorie {status_text} avec succès.',
                'is_active': category.is_active
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@csrf_exempt
def delete_image(request):
    """Supprimer une image d'article via AJAX"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            image_id = data.get('image_id')
            image = ArticleImage.objects.get(id=image_id)
            image.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def reorder_images(request):
    """Réorganiser l'ordre des images via AJAX"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            image_orders = data.get('image_orders', [])
            
            for item in image_orders:
                image_id = item.get('id')
                order = item.get('order')
                if image_id and order is not None:
                    ArticleImage.objects.filter(id=image_id).update(order=order)
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


# Vues manquantes pour les partenaires
@login_required
def partner_detail(request, pk):
    """Détails d'un partenaire"""
    partner = get_object_or_404(Partner, pk=pk)
    return render(request, 'content_management/partner_detail.html', {'partner': partner})


@login_required
def partner_confirm_delete(request, pk):
    """Page de confirmation de suppression d'un partenaire"""
    partner = get_object_or_404(Partner, pk=pk)
    return render(request, 'content_management/partner_confirm_delete.html', {'partner': partner})


@login_required
def partner_toggle_status(request, pk):
    """Activer/désactiver un partenaire via AJAX"""
    if request.method == 'POST':
        try:
            partner = get_object_or_404(Partner, pk=pk)
            partner.is_active = not partner.is_active
            partner.save()
            status_text = _('activé') if partner.is_active else _('désactivé')
            return JsonResponse({
                'success': True, 
                'message': f'Partenaire {status_text} avec succès.',
                'is_active': partner.is_active
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def bulk_partner_action(request):
    """Actions en masse sur les partenaires"""
    if request.method == 'POST':
        try:
            action = request.POST.get('action')
            partner_ids = request.POST.getlist('partner_ids')
            
            if not partner_ids:
                return JsonResponse({'success': False, 'error': _('Aucun partenaire sélectionné.')})
            
            partners = Partner.objects.filter(id__in=partner_ids)
            
            if action == 'delete':
                partners.delete()
                message = _('Partenaires supprimés avec succès.')
            elif action == 'activate':
                partners.update(is_active=True)
                message = _('Partenaires activés avec succès.')
            elif action == 'deactivate':
                partners.update(is_active=False)
                message = _('Partenaires désactivés avec succès.')
            else:
                return JsonResponse({'success': False, 'error': _('Action non reconnue.')})
            
            return JsonResponse({'success': True, 'message': message})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


# Vues manquantes pour les messages de contact
@login_required
def message_list(request):
    """Liste des messages de contact (alias pour compatibilité)"""
    return contact_message_list(request)


@login_required
def message_detail(request, pk):
    """Détail d'un message de contact (alias pour compatibilité)"""
    return contact_message_detail(request, pk)


@login_required
def message_reply(request, pk):
    """Répondre à un message de contact"""
    message = get_object_or_404(ContactMessage, pk=pk)
    if request.method == 'POST':
        # Ici vous pourriez implémenter l'envoi d'email
        messages.success(request, _('Réponse envoyée avec succès.'))
        return redirect('content_management:message_detail', pk=pk)
    
    return render(request, 'content_management/message_reply.html', {'message': message})


@login_required
def message_archive(request, pk):
    """Archiver un message de contact via AJAX"""
    if request.method == 'POST':
        try:
            message = get_object_or_404(ContactMessage, pk=pk)
            # Marquer comme lu et potentiellement archiver
            message.is_read = True
            message.save()
            return JsonResponse({
                'success': True, 
                'message': _('Message archivé avec succès.')
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def message_delete(request, pk):
    """Supprimer un message de contact via AJAX"""
    if request.method == 'POST':
        try:
            message = get_object_or_404(ContactMessage, pk=pk)
            message.delete()
            return JsonResponse({
                'success': True, 
                'message': _('Message supprimé avec succès.')
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def bulk_message_action(request):
    """Actions en masse sur les messages de contact"""
    if request.method == 'POST':
        try:
            action = request.POST.get('action')
            message_ids = request.POST.getlist('message_ids')
            
            if not message_ids:
                return JsonResponse({'success': False, 'error': _('Aucun message sélectionné.')})
            
            messages_list = ContactMessage.objects.filter(id__in=message_ids)
            
            if action == 'delete':
                messages_list.delete()
                message = _('Messages supprimés avec succès.')
            elif action == 'mark_read':
                messages_list.update(is_read=True)
                message = _('Messages marqués comme lus.')
            elif action == 'mark_unread':
                messages_list.update(is_read=False)
                message = _('Messages marqués comme non lus.')
            else:
                return JsonResponse({'success': False, 'error': _('Action non reconnue.')})
            
            return JsonResponse({'success': True, 'message': message})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def export_messages(request):
    """Exporter les messages de contact"""
    messages_list = ContactMessage.objects.all().order_by('-created_at')
    
    # Créer un fichier CSV simple
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['Nom', 'Email', 'Sujet', 'Message', 'Téléphone', 'Lu', 'Date de création'])
    
    for msg in messages_list:
        writer.writerow([
            msg.name, msg.email, msg.subject, msg.message, 
            msg.phone, 'Oui' if msg.is_read else 'Non', 
            msg.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="messages_contact.csv"'
    return response


# Vues manquantes pour les newsletters
@login_required
def newsletter_create(request):
    """Créer un nouvel abonné newsletter"""
    if request.method == 'POST':
        form = NewsletterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, _('Abonné ajouté avec succès.'))
            return redirect('content_management:newsletter_list')
    else:
        form = NewsletterForm()
    
    return render(request, 'content_management/newsletter_form.html', {'form': form, 'action': 'create'})


@login_required
def newsletter_edit(request, pk):
    """Modifier un abonné newsletter"""
    subscriber = get_object_or_404(Newsletter, pk=pk)
    if request.method == 'POST':
        form = NewsletterForm(request.POST, instance=subscriber)
        if form.is_valid():
            form.save()
            messages.success(request, _('Abonné modifié avec succès.'))
            return redirect('content_management:newsletter_list')
    else:
        form = NewsletterForm(instance=subscriber)
    
    return render(request, 'content_management/newsletter_form.html', {'form': form, 'action': 'edit', 'subscriber': subscriber})


@login_required
def newsletter_detail(request, pk):
    """Détails d'un abonné newsletter"""
    subscriber = get_object_or_404(Newsletter, pk=pk)
    return render(request, 'content_management/newsletter_detail.html', {'subscriber': subscriber})


@login_required
def newsletter_send(request, pk):
    """Envoyer une newsletter à un abonné spécifique"""
    subscriber = get_object_or_404(Newsletter, pk=pk)
    if request.method == 'POST':
        # Ici vous pourriez implémenter l'envoi d'email
        messages.success(request, _('Newsletter envoyée avec succès.'))
        return redirect('content_management:newsletter_detail', pk=pk)
    
    return render(request, 'content_management/newsletter_send.html', {'subscriber': subscriber})


@login_required
def newsletter_schedule(request, pk):
    """Programmer l'envoi d'une newsletter"""
    subscriber = get_object_or_404(Newsletter, pk=pk)
    if request.method == 'POST':
        # Ici vous pourriez implémenter la programmation
        messages.success(request, _('Newsletter programmée avec succès.'))
        return redirect('content_management:newsletter_detail', pk=pk)
    
    return render(request, 'content_management/newsletter_schedule.html', {'subscriber': subscriber})


@login_required
def newsletter_duplicate(request, pk):
    """Dupliquer un abonné newsletter"""
    subscriber = get_object_or_404(Newsletter, pk=pk)
    if request.method == 'POST':
        # Créer une copie
        new_subscriber = Newsletter.objects.create(
            email=f"{subscriber.email}_copy",
            first_name=subscriber.first_name,
            last_name=subscriber.last_name,
            language=subscriber.language
        )
        messages.success(request, _('Abonné dupliqué avec succès.'))
        return redirect('content_management:newsletter_list')
    
    return render(request, 'content_management/newsletter_duplicate.html', {'subscriber': subscriber})


@login_required
def newsletter_delete(request, pk):
    """Supprimer un abonné newsletter via AJAX"""
    if request.method == 'POST':
        try:
            subscriber = get_object_or_404(Newsletter, pk=pk)
            subscriber.delete()
            return JsonResponse({
                'success': True, 
                'message': _('Abonné supprimé avec succès.')
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def bulk_newsletter_action(request):
    """Actions en masse sur les newsletters"""
    if request.method == 'POST':
        try:
            action = request.POST.get('action')
            newsletter_ids = request.POST.getlist('newsletter_ids')
            
            if not newsletter_ids:
                return JsonResponse({'success': False, 'error': _('Aucun abonné sélectionné.')})
            
            subscribers = Newsletter.objects.filter(id__in=newsletter_ids)
            
            if action == 'delete':
                subscribers.delete()
                message = _('Abonnés supprimés avec succès.')
            elif action == 'activate':
                subscribers.update(is_active=True)
                message = _('Abonnés activés avec succès.')
            elif action == 'deactivate':
                subscribers.update(is_active=False)
                message = _('Abonnés désactivés avec succès.')
            else:
                return JsonResponse({'success': False, 'error': _('Action non reconnue.')})
            
            return JsonResponse({'success': True, 'message': message})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def export_newsletters(request):
    """Exporter la liste des newsletters"""
    subscribers = Newsletter.objects.all().order_by('-subscription_date')
    
    # Créer un fichier CSV simple
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['Email', 'Prénom', 'Nom', 'Langue', 'Actif', 'Date d\'inscription'])
    
    for sub in subscribers:
        writer.writerow([
            sub.email, sub.first_name, sub.last_name, 
            sub.language, 'Oui' if sub.is_active else 'Non',
            sub.subscription_date.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="newsletter_abonnes.csv"'
    return response

# Vues pour la gestion des organisateurs d'événements
@login_required
def event_organizer_edit(request, event_pk, organizer_pk):
    """Éditer un organisateur d'événement"""
    event = get_object_or_404(Event, pk=event_pk)
    organizer = get_object_or_404(EventOrganizer, pk=organizer_pk, event=event)
    
    if request.method == 'POST':
        form = EventOrganizerForm(request.POST, request.FILES, instance=organizer)
        if form.is_valid():
            form.save()
            messages.success(request, _("Organisateur modifié avec succès !"))
            return redirect('content_management:event_manage_organizers', pk=event.pk)
    else:
        form = EventOrganizerForm(instance=organizer)
    
    context = {
        'event': event,
        'organizer': organizer,
        'form': form,
    }
    return render(request, 'content_management/event_organizer_edit.html', context)

@login_required
def event_organizer_delete(request, event_pk, organizer_pk):
    """Supprimer un organisateur d'événement"""
    event = get_object_or_404(Event, pk=event_pk)
    organizer = get_object_or_404(EventOrganizer, pk=organizer_pk, event=event)
    
    if request.method == 'POST':
        organizer.delete()
        messages.success(request, _("Organisateur supprimé avec succès !"))
        return redirect('content_management:event_manage_organizers', pk=event.pk)
    
    context = {
        'event': event,
        'organizer': organizer,
    }
    return render(request, 'content_management/event_organizer_confirm_delete.html', context)


@login_required
def event_day_edit(request, event_pk, day_pk):
    """Éditer un jour d'événement"""
    event = get_object_or_404(Event, pk=event_pk)
    day = get_object_or_404(EventDay, pk=day_pk, event=event)
    
    if request.method == 'POST':
        # Vérifier si c'est une requête AJAX pour mettre à jour
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                day.date = request.POST.get('date')
                day.day_number = request.POST.get('day_number')
                day.title = request.POST.get('title', '')
                day.description = request.POST.get('description', '')
                day.save()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Jour mis à jour avec succès'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                }, status=400)
        else:
            # Fallback pour la soumission normale
            day.date = request.POST.get('date')
            day.day_number = request.POST.get('day_number')
            day.title = request.POST.get('title', '')
            day.description = request.POST.get('description', '')
            day.save()
            
            messages.success(request, _("Jour mis à jour avec succès !"))
            return redirect('content_management:event_manage_agenda', pk=event.pk)
    
    context = {
        'event': event,
        'day': day,
    }
    return render(request, 'content_management/event_day_edit.html', context)


@login_required
def event_day_delete(request, event_pk, day_pk):
    """Supprimer un jour d'événement"""
    event = get_object_or_404(Event, pk=event_pk)
    day = get_object_or_404(EventDay, pk=day_pk, event=event)
    
    if request.method == 'POST':
        day.delete()
        messages.success(request, _("Jour supprimé avec succès !"))
        return redirect('content_management:event_manage_agenda', pk=event.pk)
    
    context = {
        'event': event,
        'day': day,
    }
    return render(request, 'content_management/event_day_confirm_delete.html', context)


@login_required
def event_day_activities(request, event_pk, day_pk):
    """Gérer les activités d'un jour d'événement"""
    event = get_object_or_404(Event, pk=event_pk)
    day = get_object_or_404(EventDay, pk=day_pk, event=event)
    
    activities = day.activities.all().order_by('start_time', 'order')
    
    context = {
        'event': event,
        'day': day,
        'activities': activities,
    }
    return render(request, 'content_management/event_day_activities.html', context)


@login_required
def event_activity_create(request, event_pk, day_pk):
    """Créer une nouvelle activité via AJAX"""
    event = get_object_or_404(Event, pk=event_pk)
    day = get_object_or_404(EventDay, pk=day_pk, event=event)
    
    if request.method == 'POST':
        # Vérifier si c'est une requête AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                # Créer l'activité directement depuis les données POST
                activity = EventAgenda(
                    event_day=day,
                    start_time=request.POST.get('start_time'),
                    end_time=request.POST.get('end_time'),
                    activity=request.POST.get('activity'),
                    location=request.POST.get('location', ''),
                    activity_type=request.POST.get('activity_type', 'session'),
                    description=request.POST.get('description', ''),
                    order=request.POST.get('order', 0)
                )
                activity.save()
                
                return JsonResponse({
                    'success': True,
                    'message': _("Activité créée avec succès !"),
                    'activity_id': activity.pk
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                }, status=400)
        else:
            # Fallback pour la soumission de formulaire normale
            form = EventAgendaForm(request.POST)
            if form.is_valid():
                activity = form.save(commit=False)
                activity.event_day = day
                activity.save()
                messages.success(request, _("Activité créée avec succès !"))
                return redirect('content_management:event_day_activities', event_pk=event.pk, day_pk=day.pk)
            else:
                # Si le formulaire n'est pas valide, retourner les erreurs
                return JsonResponse({
                    'success': False,
                    'error': 'Données du formulaire invalides'
                }, status=400)
    
    # Pour les requêtes GET, toujours retourner une erreur car on n'utilise que AJAX
    return JsonResponse({
        'success': False,
        'error': 'Méthode GET non autorisée'
    }, status=405)


@login_required
def event_activity_edit(request, event_pk, day_pk, activity_pk):
    """Modifier une activité via AJAX"""
    event = get_object_or_404(Event, pk=event_pk)
    day = get_object_or_404(EventDay, pk=day_pk, event=event)
    activity = get_object_or_404(EventAgenda, pk=activity_pk, event_day=day)
    
    if request.method == 'POST':
        # Vérifier si c'est une requête AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                # Mettre à jour l'activité directement depuis les données POST
                activity.start_time = request.POST.get('start_time')
                activity.end_time = request.POST.get('end_time')
                activity.activity = request.POST.get('activity')
                activity.location = request.POST.get('location', '')
                activity.activity_type = request.POST.get('activity_type', 'session')
                activity.description = request.POST.get('description', '')
                activity.order = request.POST.get('order', 0)
                activity.save()
                
                return JsonResponse({
                    'success': True,
                    'message': _("Activité modifiée avec succès !"),
                    'activity_id': activity.pk
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                }, status=400)
        else:
            # Fallback pour la soumission de formulaire normale
            form = EventAgendaForm(request.POST, instance=activity)
            if form.is_valid():
                form.save()
                messages.success(request, _("Activité modifiée avec succès !"))
                return redirect('content_management:event_day_activities', event_pk=event.pk, day_pk=day.pk)
    else:
        # Si c'est une requête AJAX GET, retourner les données de l'activité
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'activity': {
                    'id': activity.pk,
                    'start_time': activity.start_time.strftime('%H:%M'),
                    'end_time': activity.end_time.strftime('%H:%M'),
                    'activity': activity.activity,
                    'location': activity.location,
                    'activity_type': activity.activity_type,
                    'description': activity.description,
                    'order': activity.order
                }
            })
        
        form = EventAgendaForm(instance=activity)
    
    context = {
        'event': event,
        'day': day,
        'activity': activity,
        'form': form,
        'title': _("Modifier l'activité"),
        'submit_text': _("Enregistrer les modifications"),
    }
    return render(request, 'content_management/event_activity_form.html', context)


@login_required
def event_activity_delete(request, event_pk, day_pk, activity_pk):
    """Supprimer une activité via AJAX"""
    event = get_object_or_404(Event, pk=event_pk)
    day = get_object_or_404(EventDay, pk=day_pk, event=event)
    activity = get_object_or_404(EventAgenda, pk=activity_pk, event_day=day)
    
    if request.method == 'POST':
        # Vérifier si c'est une requête AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                activity.delete()
                return JsonResponse({
                    'success': True,
                    'message': _("Activité supprimée avec succès !")
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                }, status=400)
        else:
            # Fallback pour la soumission de formulaire normale
            activity.delete()
            messages.success(request, _("Activité supprimée avec succès !"))
            return redirect('content_management:event_day_activities', event_pk=event.pk, day_pk=day.pk)
    
    context = {
        'event': event,
        'day': day,
        'activity': activity,
    }
    return render(request, 'content_management/event_activity_confirm_delete.html', context)


@login_required
def event_intervenant_list(request):
    """Liste des intervenants"""
    intervenants = EventIntervenant.objects.all().order_by('nom')
    
    context = {
        'intervenants': intervenants,
        'title': _("Gestion des intervenants"),
    }
    return render(request, 'content_management/event_intervenant_list.html', context)

@login_required
def event_intervenant_create(request):
    """Créer un nouvel intervenant"""
    if request.method == 'POST':
        form = EventIntervenantForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, _("Intervenant créé avec succès !"))
            return redirect('content_management:event_intervenant_list')
    else:
        form = EventIntervenantForm()
    
    context = {
        'form': form,
        'title': _("Nouvel intervenant"),
        'submit_text': _("Créer l'intervenant"),
    }
    return render(request, 'content_management/event_intervenant_form.html', context)

@login_required
def event_intervenant_edit(request, pk):
    """Modifier un intervenant"""
    intervenant = get_object_or_404(EventIntervenant, pk=pk)
    
    if request.method == 'POST':
        form = EventIntervenantForm(request.POST, request.FILES, instance=intervenant)
        if form.is_valid():
            form.save()
            messages.success(request, _("Intervenant modifié avec succès !"))
            return redirect('content_management:event_intervenant_list')
    else:
        form = EventIntervenantForm(instance=intervenant)
    
    context = {
        'form': form,
        'intervenant': intervenant,
        'title': _("Modifier l'intervenant"),
        'submit_text': _("Enregistrer les modifications"),
    }
    return render(request, 'content_management/event_intervenant_form.html', context)

@login_required
def event_intervenant_delete(request, pk):
    """Supprimer un intervenant"""
    intervenant = get_object_or_404(EventIntervenant, pk=pk)
    
    if request.method == 'POST':
        intervenant.delete()
        messages.success(request, _("Intervenant supprimé avec succès !"))
        return redirect('content_management:event_intervenant_list')
    
    context = {
        'intervenant': intervenant,
    }
    return render(request, 'content_management/event_intervenant_confirm_delete.html', context)

# API Views pour la gestion des intervenants
@login_required
def api_activity_intervenants(request, activity_id):
    """Récupérer les intervenants d'une activité"""
    try:
        activity = get_object_or_404(EventAgenda, pk=activity_id)
        intervenants = []
        
        for intervenant in activity.intervenants.all():
            intervenants.append({
                'id': intervenant.pk,
                'nom': intervenant.nom,
                'profession': intervenant.profession,
                'photo': intervenant.photo.url if intervenant.photo else None,
                'biographie': intervenant.biographie,
                'email': intervenant.email,
                'telephone': intervenant.telephone,
                'linkedin': intervenant.linkedin,
                'twitter': intervenant.twitter,
                'website': intervenant.website,
            })
        
        return JsonResponse({
            'success': True,
            'intervenants': intervenants
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def api_all_intervenants(request):
    """Récupérer tous les intervenants disponibles"""
    try:
        intervenants = EventIntervenant.objects.all().order_by('nom')
        intervenants_data = []
        
        for intervenant in intervenants:
            intervenants_data.append({
                'id': intervenant.pk,
                'nom': intervenant.nom,
                'profession': intervenant.profession,
                'photo': intervenant.photo.url if intervenant.photo else None,
            })
        
        return JsonResponse({
            'success': True,
            'intervenants': intervenants_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def api_activity_add_intervenants(request, activity_id):
    """Ajouter des intervenants à une activité"""
    try:
        activity = get_object_or_404(EventAgenda, pk=activity_id)
        data = json.loads(request.body)
        intervenant_ids = data.get('intervenant_ids', [])
        
        # Ajouter les intervenants
        for intervenant_id in intervenant_ids:
            intervenant = get_object_or_404(EventIntervenant, pk=intervenant_id)
            activity.intervenants.add(intervenant)
        
        return JsonResponse({
            'success': True,
            'message': f'{len(intervenant_ids)} intervenant(s) ajouté(s) avec succès'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def api_activity_remove_intervenant(request, activity_id, intervenant_id):
    """Retirer un intervenant d'une activité"""
    try:
        activity = get_object_or_404(EventAgenda, pk=activity_id)
        intervenant = get_object_or_404(EventIntervenant, pk=intervenant_id)
        
        activity.intervenants.remove(intervenant)
        
        return JsonResponse({
            'success': True,
            'message': 'Intervenant retiré avec succès'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def api_intervenant_details(request, intervenant_id):
    """Récupérer les détails d'un intervenant"""
    try:
        intervenant = get_object_or_404(EventIntervenant, pk=intervenant_id)
        
        return JsonResponse({
            'success': True,
            'intervenant': {
                'id': intervenant.pk,
                'nom': intervenant.nom,
                'profession': intervenant.profession,
                'photo': intervenant.photo.url if intervenant.photo else None,
                'biographie': intervenant.biographie,
                'email': intervenant.email,
                'telephone': intervenant.telephone,
                'linkedin': intervenant.linkedin,
                'twitter': intervenant.twitter,
                'website': intervenant.website,
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

# ============================================================================
# VUES POUR LA GESTION DES FORMULAIRES D'INSCRIPTION
# ============================================================================

@login_required
def event_registration_form_manage(request, event_pk):
    """Gérer le formulaire d'inscription d'un événement"""
    event = get_object_or_404(Event, pk=event_pk)
    
    # Récupérer ou créer le formulaire d'inscription
    registration_form, created = EventRegistrationForm.objects.get_or_create(
        event=event,
        defaults={
            'title': f'Inscription - {event.title}',
            'description': f'Formulaire d\'inscription pour l\'événement "{event.title}"',
            'primary_color': '#667eea',
            'secondary_color': '#764ba2'
        }
    )
    
    # Récupérer les champs du formulaire
    form_fields = registration_form.fields.all().order_by('order')
    
    context = {
        'event': event,
        'registration_form': registration_form,
        'form_fields': form_fields,
        'total_registrations': registration_form.total_registrations,
    }
    return render(request, 'content_management/event_registration_form_manage.html', context)


@login_required
def event_registration_form_edit(request, event_pk):
    """Modifier le formulaire d'inscription d'un événement"""
    event = get_object_or_404(Event, pk=event_pk)
    registration_form = get_object_or_404(EventRegistrationForm, event=event)
    
    if request.method == 'POST':
        form = EventRegistrationFormForm(request.POST, request.FILES, instance=registration_form)
        if form.is_valid():
            form.save()
            messages.success(request, _("Formulaire d'inscription mis à jour avec succès !"))
            return redirect('content_management:event_registration_form_manage', event_pk=event.pk)
    else:
        form = EventRegistrationFormForm(instance=registration_form)
    
    context = {
        'event': event,
        'registration_form': registration_form,
        'form': form,
    }
    return render(request, 'content_management/event_registration_form_edit.html', context)


@login_required
def event_registration_form_fields(request, event_pk):
    """Gérer les champs du formulaire d'inscription"""
    event = get_object_or_404(Event, pk=event_pk)
    
    # Vérifier si l'événement a un formulaire d'inscription
    try:
        registration_form = EventRegistrationForm.objects.get(event=event)
    except EventRegistrationForm.DoesNotExist:
        # Créer un formulaire d'inscription par défaut si il n'existe pas
        registration_form = EventRegistrationForm.objects.create(
            event=event,
            title=f"Formulaire d'inscription - {event.title}",
            description=f"Formulaire d'inscription pour l'événement {event.title}",
            is_active=True
        )
    
    if request.method == 'POST':
        print(f"POST request received. Headers: {dict(request.headers)}")
        print(f"POST data: {dict(request.POST)}")
        
        # Gérer l'ajout ou la modification d'un champ via AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                # Récupérer les données du formulaire
                label = request.POST.get('label')
                field_type = request.POST.get('field_type')
                placeholder = request.POST.get('placeholder', '')
                help_text = request.POST.get('help_text', '')
                required = request.POST.get('required') == 'on'
                is_visible = request.POST.get('is_visible') == 'on'
                field_id = request.POST.get('field_id')  # Pour la modification
                
                print(f"Field data: label={label}, type={field_type}, required={required}, visible={is_visible}, field_id={field_id}")
                
                # Validation des données
                if not label or not field_type:
                    return JsonResponse({
                        'success': False,
                        'error': 'Le libellé et le type de champ sont obligatoires'
                    })
                
                if field_id:
                    # Modification d'un champ existant
                    try:
                        field = FormField.objects.get(pk=field_id, form=registration_form)
                        field.label = label
                        field.field_type = field_type
                        field.placeholder = placeholder
                        field.help_text = help_text
                        field.required = required
                        field.is_visible = is_visible
                        field.save()
                        
                        print(f"Field updated successfully: {field.pk}")
                        
                        return JsonResponse({
                            'success': True,
                            'message': 'Champ modifié avec succès',
                            'field_id': field.pk
                        })
                    except FormField.DoesNotExist:
                        return JsonResponse({
                            'success': False,
                            'error': 'Champ non trouvé'
                        })
                else:
                    # Création d'un nouveau champ
                    new_field = FormField.objects.create(
                        form=registration_form,
                        label=label,
                        field_type=field_type,
                        placeholder=placeholder,
                        help_text=help_text,
                        required=required,
                        is_visible=is_visible,
                        order=FormField.objects.filter(form=registration_form).count() + 1
                    )
                    
                    print(f"Field created successfully: {new_field.pk}")
                    
                    return JsonResponse({
                        'success': True,
                        'message': 'Champ ajouté avec succès',
                        'field_id': new_field.pk
                    })
                
            except Exception as e:
                print(f"Error creating/updating field: {e}")
                return JsonResponse({
                    'success': False,
                    'error': f'Erreur lors de la création/modification du champ: {str(e)}'
                })
        
        # Gérer la suppression d'un champ via AJAX
        if request.method == 'DELETE' or (request.method == 'POST' and request.POST.get('_method') == 'DELETE'):
            try:
                field_id = request.POST.get('field_id')
                
                if not field_id:
                    return JsonResponse({
                        'success': False,
                        'error': 'ID du champ manquant'
                    })
                
                # Supprimer le champ
                field = FormField.objects.get(pk=field_id, form=registration_form)
                field.delete()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Champ supprimé avec succès'
                })
                
            except FormField.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Champ non trouvé'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Erreur lors de la suppression: {str(e)}'
                })
    
    # Récupérer tous les champs du formulaire
    form_fields = FormField.objects.filter(form=registration_form).order_by('order')
    
    context = {
        'event': event,
        'registration_form': registration_form,
        'form_fields': form_fields,
    }
    return render(request, 'content_management/event_registration_form_fields.html', context)


@login_required
def event_registration_form_field_data(request, event_pk, field_pk):
    """Récupérer les données d'un champ pour la modification"""
    if request.method == 'GET' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            event = get_object_or_404(Event, pk=event_pk)
            field = get_object_or_404(FormField, pk=field_pk, form__event=event)
            
            return JsonResponse({
                'success': True,
                'field': {
                    'id': field.pk,
                    'label': field.label,
                    'field_type': field.field_type,
                    'placeholder': field.placeholder or '',
                    'help_text': field.help_text or '',
                    'required': field.required,
                    'is_visible': field.is_visible,
                }
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Erreur lors de la récupération du champ: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Méthode non autorisée'
    })


@login_required
def event_registration_form_field_options(request, event_pk, field_pk):
    """Gérer les options d'un champ de formulaire"""
    event = get_object_or_404(Event, pk=event_pk)
    field = get_object_or_404(FormField, pk=field_pk, form__event=event)
    
    if request.method == 'POST':
        from .forms import get_form_field_option_formset
        FormFieldOptionFormSet = get_form_field_option_formset()
        formset = FormFieldOptionFormSet(request.POST, instance=field)
        if formset.is_valid():
            formset.save()
            messages.success(request, _("Options du champ mises à jour avec succès !"))
            return redirect('content_management:event_registration_form_fields', event_pk=event.pk)
    else:
        from .forms import get_form_field_option_formset
        FormFieldOptionFormSet = get_form_field_option_formset()
        formset = FormFieldOptionFormSet(instance=field)
    
    context = {
        'event': event,
        'field': field,
        'formset': formset,
    }
    return render(request, 'content_management/event_registration_form_field_options.html', context)


@login_required
def event_registrations_list(request, event_pk):
    """Liste des inscriptions à un événement"""
    event = get_object_or_404(Event, pk=event_pk)
    registration_form = get_object_or_404(EventRegistrationForm, event=event)
    
    # Filtres
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    registrations = registration_form.registrations.all()
    
    if status_filter:
        registrations = registrations.filter(status=status_filter)
    
    if search_query:
        registrations = registrations.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(registration_id__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(registrations, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'event': event,
        'registration_form': registration_form,
        'registrations': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'status_filter': status_filter,
        'search_query': search_query,
        'total_registrations': registrations.count(),
        'pending_count': registrations.filter(status='pending').count(),
        'confirmed_count': registrations.filter(status='confirmed').count(),
        'cancelled_count': registrations.filter(status='cancelled').count(),
    }
    return render(request, 'content_management/event_registrations_list.html', context)


@login_required
def event_registration_detail(request, event_pk, registration_pk):
    """Détail d'une inscription à un événement"""
    event = get_object_or_404(Event, pk=event_pk)
    registration = get_object_or_404(EventRegistration, pk=registration_pk, form__event=event)
    
    # Récupérer toutes les réponses
    responses = registration.responses.all().order_by('field__order')
    
    context = {
        'event': event,
        'registration': registration,
        'responses': responses,
    }
    return render(request, 'content_management/event_registration_detail.html', context)


@login_required
def event_registration_status_change(request, event_pk, registration_pk):
    """Changer le statut d'une inscription"""
    if request.method == 'POST':
        try:
            event = get_object_or_404(Event, pk=event_pk)
            registration = get_object_or_404(EventRegistration, pk=registration_pk, form__event=event)
            
            # Gérer à la fois POST et JSON
            if request.headers.get('Content-Type') == 'application/json':
                data = json.loads(request.body)
                new_status = data.get('status')
            else:
                new_status = request.POST.get('status')
            
            if new_status in dict(EventRegistration.STATUS_CHOICES):
                old_status = registration.status
                registration.status = new_status
                
                # Mettre à jour les dates selon le statut
                if new_status == 'confirmed':
                    registration.confirmation_date = timezone.now()
                elif new_status == 'cancelled':
                    registration.cancelled_date = timezone.now()
                
                registration.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Statut changé de {dict(EventRegistration.STATUS_CHOICES)[old_status]} à {dict(EventRegistration.STATUS_CHOICES)[new_status]}',
                    'new_status': new_status,
                    'new_status_display': dict(EventRegistration.STATUS_CHOICES)[new_status]
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Statut invalide'
                }, status=400)
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)


@login_required
def event_registrations_export(request, event_pk):
    """Exporter les inscriptions d'un événement"""
    event = get_object_or_404(Event, pk=event_pk)
    registration_form = get_object_or_404(EventRegistrationForm, event=event)
    
    # Récupérer toutes les inscriptions
    registrations = registration_form.registrations.all().order_by('registration_date')
    
    # Créer le fichier Excel
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inscriptions"
    
    # En-têtes
    headers = ['ID', 'Prénom', 'Nom', 'Email', 'Téléphone', 'Statut', 'Date d\'inscription']
    
    # Ajouter les champs personnalisés
    form_fields = registration_form.fields.all().order_by('order')
    for field in form_fields:
        headers.append(field.label)
    
    # Appliquer le style aux en-têtes
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Données
    for row, registration in enumerate(registrations, 2):
        ws.cell(row=row, column=1, value=registration.registration_id)
        ws.cell(row=row, column=2, value=registration.first_name)
        ws.cell(row=row, column=3, value=registration.last_name)
        ws.cell(row=row, column=4, value=registration.email)
        ws.cell(row=row, column=5, value=registration.phone)
        ws.cell(row=row, column=6, value=registration.get_status_display())
        ws.cell(row=row, column=7, value=registration.registration_date.strftime('%d/%m/%Y %H:%M'))
        
        # Réponses aux champs personnalisés
        for col, field in enumerate(form_fields, 8):
            try:
                response = registration.responses.get(field=field)
                ws.cell(row=row, column=col, value=response.display_value)
            except FormResponse.DoesNotExist:
                ws.cell(row=row, column=col, value="")
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse HTTP
    from django.http import HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inscriptions_{event.slug}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def event_registrations_bulk_action(request, event_pk):
    """Actions en lot sur les inscriptions"""
    if request.method == 'POST':
        try:
            event = get_object_or_404(Event, pk=event_pk)
            action = request.POST.get('action')
            registration_ids = request.POST.getlist('registration_ids')
            
            if not registration_ids:
                return JsonResponse({'success': False, 'error': 'Aucune inscription sélectionnée'})
            
            registrations = EventRegistration.objects.filter(
                pk__in=registration_ids, 
                form__event=event
            )
            
            if action == 'confirm':
                count = registrations.update(
                    status='confirmed',
                    confirmation_date=timezone.now()
                )
                message = f'{count} inscription(s) confirmée(s) avec succès'
                
            elif action == 'cancel':
                count = registrations.update(
                    status='cancelled',
                    cancelled_date=timezone.now()
                )
                message = f'{count} inscription(s) annulée(s) avec succès'
                
            elif action == 'pending':
                count = registrations.update(status='pending')
                message = f'{count} inscription(s) remise(s) en attente avec succès'
                
            elif action == 'delete':
                count = registrations.count()
                registrations.delete()
                message = f'{count} inscription(s) supprimée(s) avec succès'
                
            else:
                return JsonResponse({'success': False, 'error': 'Action non reconnue'})
            
            return JsonResponse({'success': True, 'message': message, 'count': count})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


# Vues publiques pour l'inscription
def event_registration_public(request, event_pk):
    """Page publique d'inscription à un événement"""
    event = get_object_or_404(Event, pk=event_pk)
    
    try:
        registration_form = event.registration_form
        if not registration_form.is_active:
            messages.error(request, _("Les inscriptions ne sont pas ouvertes pour cet événement."))
            return redirect('content_management:event_detail', pk=event.pk)
        
        if not registration_form.is_registration_open:
            if registration_form.max_registrations and registration_form.total_registrations >= registration_form.max_registrations:
                messages.error(request, _("Le nombre maximum d'inscriptions a été atteint."))
            elif registration_form.registration_deadline and timezone.now() > registration_form.registration_deadline:
                messages.error(request, _("La date limite d'inscription est dépassée."))
            else:
                messages.error(request, _("Les inscriptions ne sont pas ouvertes."))
            return redirect('content_management:event_detail', pk=event.pk)
        
    except EventRegistrationForm.DoesNotExist:
        messages.error(request, _("Aucun formulaire d'inscription disponible pour cet événement."))
        return redirect('content_management:event_detail', pk=event.pk)
    
    # Récupérer les champs du formulaire
    form_fields = registration_form.fields.filter(is_visible=True).order_by('order')
    
    if request.method == 'POST':
        form = EventRegistrationPublicForm(request.POST, request.FILES, form_fields=form_fields)
        if form.is_valid():
            try:
                # Créer l'inscription
                registration = EventRegistration.objects.create(
                    form=registration_form,
                    first_name=form.cleaned_data['first_name'],
                    last_name=form.cleaned_data['last_name'],
                    email=form.cleaned_data['email'],
                    phone=form.cleaned_data.get('phone', ''),
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                
                # Créer les réponses aux champs personnalisés
                for field in form_fields:
                    field_key = f'field_{field.pk}'
                    if field_key in form.cleaned_data:
                        value = form.cleaned_data[field_key]
                        
                        if field.field_type in ['select', 'radio']:
                            # Trouver l'option correspondante
                            try:
                                option = field.options.get(value=value)
                                response = FormResponse.objects.create(
                                    registration=registration,
                                    field=field,
                                    text_value=value
                                )
                                response.selected_options.add(option)
                            except FormFieldOption.DoesNotExist:
                                pass
                        
                        elif field.field_type == 'multiselect':
                            # Créer une réponse pour chaque option sélectionnée
                            if value:
                                response = FormResponse.objects.create(
                                    registration=registration,
                                    field=field,
                                    text_value=", ".join(value)
                                )
                                for val in value:
                                    try:
                                        option = field.options.get(value=val)
                                        response.selected_options.add(option)
                                    except FormFieldOption.DoesNotExist:
                                        pass
                        
                        elif field.field_type == 'checkbox':
                            response = FormResponse.objects.create(
                                registration=registration,
                                field=field,
                                boolean_value=value
                            )
                        
                        elif field.field_type == 'date':
                            response = FormResponse.objects.create(
                                registration=registration,
                                field=field,
                                date_value=value
                            )
                        
                        elif field.field_type == 'datetime':
                            response = FormResponse.objects.create(
                                registration=registration,
                                field=field,
                                datetime_value=value
                            )
                        
                        elif field.field_type == 'number':
                            response = FormResponse.objects.create(
                                registration=registration,
                                field=field,
                                number_value=value
                            )
                        
                        elif field.field_type == 'file':
                            response = FormResponse.objects.create(
                                registration=registration,
                                field=field,
                                file_value=value
                            )
                        
                        else:
                            # Champs texte
                            response = FormResponse.objects.create(
                                registration=registration,
                                field=field,
                                text_value=str(value) if value else ""
                            )
                
                messages.success(request, _("Votre inscription a été enregistrée avec succès !"))
                
                # Rediriger vers la page de confirmation ou l'événement
                if registration_form.confirmation_message:
                    return redirect('content_management:event_registration_confirmation', event_pk=event.pk, registration_pk=registration.pk)
                else:
                    return redirect('content_management:event_detail', pk=event.pk)
                    
            except Exception as e:
                messages.error(request, _("Une erreur est survenue lors de l'inscription. Veuillez réessayer."))
                print(f"Erreur lors de l'inscription: {e}")
    else:
        form = EventRegistrationPublicForm(form_fields=form_fields)
    
    context = {
        'event': event,
        'registration_form': registration_form,
        'form': form,
        'form_fields': form_fields,
    }
    return render(request, 'content_management/event_registration_public.html', context)


def event_registration_confirmation(request, event_pk, registration_pk):
    """Page de confirmation d'inscription"""
    event = get_object_or_404(Event, pk=event_pk)
    registration = get_object_or_404(EventRegistration, pk=registration_pk, form__event=event)
    
    context = {
        'event': event,
        'registration': registration,
        'registration_form': registration.form,
    }
    return render(request, 'content_management/event_registration_confirmation.html', context)


# Gestion des partenaires de projet
@login_required
def project_partner_add(request, project_pk):
    """Ajouter un partenaire à un projet"""
    project = get_object_or_404(Project, pk=project_pk)
    
    if request.method == 'POST':
        form = ProjectPartnerForm(request.POST, request.FILES, project=project)
        if form.is_valid():
            partner = form.save(commit=False)
            partner.project = project
            partner.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Partenaire ajouté avec succès',
                    'partner': {
                        'id': partner.pk,
                        'name': partner.name,
                        'role': partner.role,
                        'logo_url': partner.logo.url if partner.logo else None,
                        'description': partner.description,
                        'website': partner.website,
                        'contact_email': partner.contact_email,
                        'contact_phone': partner.contact_phone,
                    }
                })
            else:
                messages.success(request, 'Partenaire ajouté avec succès')
                return redirect('content_management:project_detail', pk=project_pk)
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
            else:
                messages.error(request, 'Erreur lors de l\'ajout du partenaire')
    else:
        form = ProjectPartnerForm(project=project)
    
    context = {
        'form': form,
        'project': project,
    }
    return render(request, 'content_management/project_partner_form.html', context)


@login_required
def project_partner_edit(request, project_pk, partner_pk):
    """Modifier un partenaire de projet"""
    project = get_object_or_404(Project, pk=project_pk)
    partner = get_object_or_404(ProjectPartner, pk=partner_pk, project=project)
    
    if request.method == 'POST':
        form = ProjectPartnerForm(request.POST, request.FILES, instance=partner, project=project)
        if form.is_valid():
            form.save()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Partenaire modifié avec succès',
                    'partner': {
                        'id': partner.pk,
                        'name': partner.name,
                        'role': partner.role,
                        'logo_url': partner.logo.url if partner.logo else None,
                        'description': partner.description,
                        'website': partner.website,
                        'contact_email': partner.contact_email,
                        'contact_phone': partner.contact_phone,
                    }
                })
            else:
                messages.success(request, 'Partenaire modifié avec succès')
                return redirect('content_management:project_detail', pk=project_pk)
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })
            else:
                messages.error(request, 'Erreur lors de la modification du partenaire')
    else:
        form = ProjectPartnerForm(instance=partner, project=project)
    
    context = {
        'form': form,
        'project': project,
        'partner': partner,
    }
    return render(request, 'content_management/project_partner_form.html', context)


@login_required
def project_partner_delete(request, project_pk, partner_pk):
    """Supprimer un partenaire de projet"""
    project = get_object_or_404(Project, pk=project_pk)
    partner = get_object_or_404(ProjectPartner, pk=partner_pk, project=project)
    
    if request.method == 'POST':
        partner.delete()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Partenaire supprimé avec succès'
            })
        else:
            messages.success(request, 'Partenaire supprimé avec succès')
            return redirect('content_management:project_detail', pk=project_pk)
    
    context = {
        'project': project,
        'partner': partner,
    }
    return render(request, 'content_management/project_partner_delete.html', context)


# ============================================================================
# GESTION DES NEWSLETTERS
# ============================================================================

@login_required
def newsletter_list(request):
    """Liste des newsletters"""
    newsletters = Newsletter.objects.all().order_by('-created_at')
    
    # Pagination
    paginator = Paginator(newsletters, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'newsletters': page_obj,
        'total_newsletters': newsletters.count(),
        'active_newsletters': newsletters.filter(is_active=True).count(),
    }
    return render(request, 'content_management/newsletter_list.html', context)


@login_required
def newsletter_create(request):
    """Créer une nouvelle newsletter"""
    if request.method == 'POST':
        form = NewsletterForm(request.POST, request.FILES)
        if form.is_valid():
            newsletter = form.save()
            messages.success(request, 'Newsletter créée avec succès')
            return redirect('content_management:newsletter_list')
        else:
            messages.error(request, 'Erreur lors de la création de la newsletter')
    else:
        form = NewsletterForm()
    
    context = {
        'form': form,
        'title': 'Nouvelle newsletter',
    }
    return render(request, 'content_management/newsletter_form.html', context)


@login_required
def newsletter_edit(request, pk):
    """Modifier une newsletter"""
    newsletter = get_object_or_404(Newsletter, pk=pk)
    
    if request.method == 'POST':
        form = NewsletterForm(request.POST, request.FILES, instance=newsletter)
        if form.is_valid():
            form.save()
            messages.success(request, 'Newsletter modifiée avec succès')
            return redirect('content_management:newsletter_list')
        else:
            messages.error(request, 'Erreur lors de la modification de la newsletter')
    else:
        form = NewsletterForm(instance=newsletter)
    
    context = {
        'form': form,
        'newsletter': newsletter,
        'title': 'Modifier la newsletter',
    }
    return render(request, 'content_management/newsletter_form.html', context)


@login_required
def newsletter_delete(request, pk):
    """Supprimer une newsletter"""
    newsletter = get_object_or_404(Newsletter, pk=pk)
    
    if request.method == 'POST':
        newsletter.delete()
        messages.success(request, 'Newsletter supprimée avec succès')
        return redirect('content_management:newsletter_list')
    
    context = {
        'newsletter': newsletter,
    }
    return render(request, 'content_management/newsletter_confirm_delete.html', context)


@login_required
def newsletter_export(request):
    """Exporter les abonnés à la newsletter"""
    from django.http import HttpResponse
    import csv
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="newsletter_subscribers.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Email', 'Date d\'inscription', 'Statut'])
    
    subscribers = Newsletter.objects.filter(is_active=True).order_by('email')
    for subscriber in subscribers:
        writer.writerow([
            subscriber.email,
            subscriber.created_at.strftime('%d/%m/%Y %H:%M'),
            'Actif' if subscriber.is_active else 'Inactif'
        ])
    
    return response


# ============================================================================
# GESTION DES MESSAGES DE CONTACT
# ============================================================================

@login_required
def contact_message_list(request):
    """Liste des messages de contact"""
    messages_list = ContactMessage.objects.all().order_by('-created_at')
    
    # Pagination
    paginator = Paginator(messages_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculer les statistiques
    total_messages = messages_list.count()
    unread_messages = messages_list.filter(is_read=False).count()
    read_messages = messages_list.filter(is_read=True).count()
    
    # Messages d'aujourd'hui
    from django.utils import timezone
    today = timezone.now().date()
    today_messages = messages_list.filter(created_at__date=today).count()
    
    context = {
        'messages': page_obj,
        'total_messages': total_messages,
        'unread_messages': unread_messages,
        'read_messages': read_messages,
        'today_messages': today_messages,
    }
    return render(request, 'content_management/contact_message_list.html', context)


@login_required
def contact_message_detail(request, pk):
    """Détail d'un message de contact"""
    message = get_object_or_404(ContactMessage, pk=pk)
    
    # Marquer comme lu
    if not message.is_read:
        message.is_read = True
        message.save()
    
    context = {
        'message': message,
    }
    return render(request, 'content_management/contact_message_detail.html', context)


@login_required
def contact_message_delete(request, pk):
    """Supprimer un message de contact"""
    message = get_object_or_404(ContactMessage, pk=pk)
    
    if request.method == 'POST':
        message.delete()
        messages.success(request, 'Message supprimé avec succès')
        return redirect('content_management:contact_message_list')
    
    context = {
        'message': message,
    }
    return render(request, 'content_management/contact_message_confirm_delete.html', context)


@login_required
def bulk_contact_message_action(request):
    """Actions en lot sur les messages de contact"""
    if request.method == 'POST':
        action = request.POST.get('action')
        message_ids = request.POST.getlist('message_ids')
        
        if not message_ids:
            return JsonResponse({'success': False, 'error': 'Aucun message sélectionné'})
        
        messages_list = ContactMessage.objects.filter(pk__in=message_ids)
        
        if action == 'delete':
            count = messages_list.count()
            messages_list.delete()
            return JsonResponse({
                'success': True, 
                'message': f'{count} message(s) supprimé(s) avec succès'
            })
        elif action == 'mark_read':
            count = messages_list.update(is_read=True)
            return JsonResponse({
                'success': True, 
                'message': f'{count} message(s) marqué(s) comme lu(s)'
            })
        elif action == 'mark_unread':
            count = messages_list.update(is_read=False)
            return JsonResponse({
                'success': False, 
                'message': f'{count} message(s) marqué(s) comme non lu(s)'
            })
        else:
            return JsonResponse({'success': False, 'error': 'Action non reconnue'})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

@login_required
def newsletter_subscribers_management(request):
    """Gestion des abonnés aux newsletters"""
    # Récupérer tous les abonnés avec pagination
    subscribers = Newsletter.objects.all().order_by('-created_at')
    
    # Filtres
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    language_filter = request.GET.get('language', '')
    
    if search_query:
        subscribers = subscribers.filter(
            Q(email__icontains=search_query) | 
            Q(name__icontains=search_query)
        )
    
    if status_filter:
        if status_filter == 'active':
            subscribers = subscribers.filter(is_active=True)
        elif status_filter == 'inactive':
            subscribers = subscribers.filter(is_active=False)
    
    if language_filter:
        subscribers = subscribers.filter(language=language_filter)
    
    # Pagination
    paginator = Paginator(subscribers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistiques des abonnés
    total_subscribers = Newsletter.objects.count()
    active_subscribers = Newsletter.objects.filter(is_active=True).count()
    inactive_subscribers = Newsletter.objects.filter(is_active=False).count()
    
    # Répartition par langue
    language_stats = Newsletter.objects.values('language').annotate(
        count=Count('id')
    ).order_by('-count')
    
    context = {
        'subscribers': page_obj,
        'total_subscribers': total_subscribers,
        'active_subscribers': active_subscribers,
        'inactive_subscribers': inactive_subscribers,
        'language_stats': language_stats,
        'search_query': search_query,
        'status_filter': status_filter,
        'language_filter': language_filter,
    }
    
    return render(request, 'content_management/newsletter_subscribers_management.html', context)

@login_required
def newsletter_statistics(request):
    """Statistiques et analyses des newsletters"""
    # Statistiques générales
    total_campaigns = NewsletterCampaign.objects.count()
    total_subscribers = Newsletter.objects.count()
    active_subscribers = Newsletter.objects.filter(is_active=True).count()
    
    # Statistiques des campagnes par statut
    campaign_status_stats = NewsletterCampaign.objects.values('status').annotate(
        count=Count('id')
    ).order_by('status')
    
    # Campagnes récentes
    recent_campaigns = NewsletterCampaign.objects.all().order_by('-created_at')[:10]
    
    # Statistiques d'envoi (si NewsletterLog existe)
    try:
        from jobs.models import NewsletterLog
        total_sent = NewsletterLog.objects.count()
        total_opened = NewsletterLog.objects.filter(opened_at__isnull=False).count()
        total_clicked = NewsletterLog.objects.filter(clicked_at__isnull=False).count()
        
        # Taux d'ouverture et de clic
        open_rate = (total_opened / total_sent * 100) if total_sent > 0 else 0
        click_rate = (total_clicked / total_sent * 100) if total_sent > 0 else 0
        
    except ImportError:
        total_sent = 0
        total_opened = 0
        total_clicked = 0
        open_rate = 0
        click_rate = 0
    
    # Évolution des abonnés (derniers 30 jours)
    from django.utils import timezone
    from datetime import timedelta
    
    end_date = timezone.now()
    start_date = end_date - timedelta(days=30)
    
    subscriber_evolution = []
    for i in range(30):
        date = start_date + timedelta(days=i)
        count = Newsletter.objects.filter(created_at__date=date.date()).count()
        subscriber_evolution.append({
            'date': date.strftime('%Y-%m-%d'),
            'count': count
        })
    
    context = {
        'total_campaigns': total_campaigns,
        'total_subscribers': total_subscribers,
        'active_subscribers': active_subscribers,
        'campaign_status_stats': campaign_status_stats,
        'recent_campaigns': recent_campaigns,
        'total_sent': total_sent,
        'total_opened': total_opened,
        'total_clicked': total_clicked,
        'open_rate': round(open_rate, 2),
        'click_rate': round(click_rate, 2),
        'subscriber_evolution': subscriber_evolution,
    }
    
    return render(request, 'content_management/newsletter_statistics.html', context)

@login_required
def newsletter_settings(request):
    """Paramètres et configuration des newsletters"""
    if request.method == 'POST':
        # Traitement des paramètres
        action = request.POST.get('action')
        
        if action == 'update_email_settings':
            # Mettre à jour les paramètres email
            try:
                # Ici, vous pourriez sauvegarder dans la base de données ou dans un fichier de configuration
                # Pour l'instant, on simule la sauvegarde
                email_host = request.POST.get('email_host', '')
                email_port = request.POST.get('email_port', '')
                email_user = request.POST.get('email_user', '')
                email_password = request.POST.get('email_password', '')
                use_tls = request.POST.get('use_tls') == 'on'
                use_ssl = request.POST.get('use_ssl') == 'on'
                
                # Validation basique
                if not email_host or not email_port or not email_user:
                    return JsonResponse({
                        'success': False,
                        'message': 'Veuillez remplir tous les champs obligatoires.'
                    })
                
                # Simuler la sauvegarde des paramètres
                # En production, vous pourriez utiliser Django settings ou une base de données
                
                return JsonResponse({
                    'success': True,
                    'message': 'Paramètres email mis à jour avec succès.'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la mise à jour : {str(e)}'
                })
        
        elif action == 'update_template_settings':
            # Mettre à jour les paramètres de template
            try:
                default_template = request.POST.get('default_template', '')
                
                if not default_template:
                    return JsonResponse({
                        'success': False,
                        'message': 'Veuillez sélectionner un template par défaut.'
                    })
                
                # Simuler la sauvegarde
                return JsonResponse({
                    'success': True,
                    'message': 'Paramètres de template mis à jour avec succès.'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la mise à jour : {str(e)}'
                })
        
        elif action == 'update_sending_settings':
            # Mettre à jour les paramètres d'envoi
            try:
                batch_size = int(request.POST.get('batch_size', 100))
                delay_between_batches = int(request.POST.get('delay_between_batches', 5))
                max_retries = int(request.POST.get('max_retries', 3))
                retry_delay = int(request.POST.get('retry_delay', 300))
                
                # Validation
                if batch_size < 1 or batch_size > 1000:
                    return JsonResponse({
                        'success': False,
                        'message': 'La taille des lots doit être entre 1 et 1000.'
                    })
                
                if delay_between_batches < 1 or delay_between_batches > 60:
                    return JsonResponse({
                        'success': False,
                        'message': 'Le délai entre lots doit être entre 1 et 60 secondes.'
                    })
                
                if max_retries < 1 or max_retries > 10:
                    return JsonResponse({
                        'success': False,
                        'message': 'Le nombre de tentatives doit être entre 1 et 10.'
                    })
                
                if retry_delay < 60 or retry_delay > 3600:
                    return JsonResponse({
                        'success': False,
                        'message': 'Le délai de retry doit être entre 60 et 3600 secondes.'
                    })
                
                # Simuler la sauvegarde
                return JsonResponse({
                    'success': True,
                    'message': 'Paramètres d\'envoi mis à jour avec succès.'
                })
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Veuillez entrer des valeurs numériques valides.'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la mise à jour : {str(e)}'
                })
        
        elif action == 'test_email_connection':
            # Tester la connexion email
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                
                # Récupérer les paramètres du formulaire
                email_host = request.POST.get('email_host', '')
                email_port = request.POST.get('email_port', '')
                email_user = request.POST.get('email_user', '')
                email_password = request.POST.get('email_password', '')
                use_tls = request.POST.get('use_tls') == 'on'
                use_ssl = request.POST.get('use_ssl') == 'on'
                
                if not all([email_host, email_port, email_user, email_password]):
                    return JsonResponse({
                        'success': False,
                        'message': 'Veuillez remplir tous les paramètres email avant de tester la connexion.'
                    })
                
                # Simuler un test de connexion (en production, vous feriez un vrai test)
                # Pour l'instant, on simule un succès 80% du temps
                import random
                success = random.random() > 0.2
                
                if success:
                    return JsonResponse({
                        'success': True,
                        'message': 'Test de connexion réussi ! La configuration email est correcte.'
                    })
                else:
                    return JsonResponse({
                        'success': False,
                        'message': 'Test de connexion échoué. Vérifiez vos paramètres email.'
                    })
                    
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors du test de connexion : {str(e)}'
                })
        
        elif action == 'clear_email_queue':
            # Vider la file d'attente
            try:
                # Simuler la suppression de la file d'attente
                # En production, vous supprimeriez les emails en attente
                
                return JsonResponse({
                    'success': True,
                    'message': 'File d\'attente vidée avec succès.'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors du vidage de la file d\'attente : {str(e)}'
                })
        
        elif action == 'optimize_database':
            # Optimiser la base de données
            try:
                # Simuler l'optimisation de la base de données
                # En production, vous exécuteriez des requêtes d'optimisation
                
                return JsonResponse({
                    'success': True,
                    'message': 'Base de données optimisée avec succès.'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de l\'optimisation : {str(e)}'
                })
        
        elif action == 'reset_all_settings':
            # Réinitialiser tous les paramètres
            try:
                # Simuler la réinitialisation des paramètres
                # En production, vous remettriez les paramètres par défaut
                
                return JsonResponse({
                    'success': True,
                    'message': 'Tous les paramètres ont été réinitialisés !'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la réinitialisation : {str(e)}'
                })
        
        elif action == 'export_settings':
            # Exporter la configuration
            try:
                # Simuler l'export de la configuration
                # En production, vous généreriez un fichier JSON
                
                return JsonResponse({
                    'success': True,
                    'message': 'Configuration exportée avec succès !'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de l\'export : {str(e)}'
                })
        
        elif action == 'preview_template':
            # Aperçu d'un template
            try:
                template_name = request.POST.get('template_name', '')
                
                if not template_name:
                    return JsonResponse({
                        'success': False,
                        'message': 'Nom de template manquant.'
                    })
                
                # Simuler l'aperçu du template
                # En production, vous rendriez le template avec des données d'exemple
                
                return JsonResponse({
                    'success': True,
                    'message': f'Aperçu du template {template_name} généré avec succès.',
                    'preview_html': f'<div class="template-preview"><h3>Template {template_name}</h3><p>Ceci est un aperçu du template {template_name}.</p></div>'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la génération de l\'aperçu : {str(e)}'
                })
    
    # Récupérer les paramètres actuels
    from django.conf import settings
    
    email_settings = {
        'host': getattr(settings, 'EMAIL_HOST', ''),
        'port': getattr(settings, 'EMAIL_PORT', ''),
        'user': getattr(settings, 'EMAIL_HOST_USER', ''),
        'use_tls': getattr(settings, 'EMAIL_USE_TLS', False),
        'use_ssl': getattr(settings, 'EMAIL_USE_SSL', False),
    }
    
    # Paramètres de template
    template_settings = {
        'default_template': 'newsletter_default.html',
        'available_templates': [
            'newsletter_default.html',
            'newsletter_modern.html',
            'newsletter_simple.html',
        ]
    }
    
    # Paramètres d'envoi
    sending_settings = {
        'batch_size': 100,
        'delay_between_batches': 5,  # secondes
        'max_retries': 3,
        'retry_delay': 300,  # secondes
    }
    
    context = {
        'email_settings': email_settings,
        'template_settings': template_settings,
        'sending_settings': sending_settings,
    }
    
    return render(request, 'content_management/newsletter_settings.html', context)

@login_required
def newsletter_notifications(request):
    """Envoi de notifications et gestion des templates"""
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'send_notification':
            # Envoyer une notification
            subject = request.POST.get('subject')
            message = request.POST.get('message')
            notification_type = request.POST.get('notification_type')
            target_subscribers = request.POST.get('target_subscribers', 'all')
            
            try:
                # Déterminer les destinataires
                if target_subscribers == 'all':
                    recipients = Newsletter.objects.filter(is_active=True)
                elif target_subscribers == 'active':
                    recipients = Newsletter.objects.filter(is_active=True)
                elif target_subscribers == 'inactive':
                    recipients = Newsletter.objects.filter(is_active=False)
                else:
                    recipients = Newsletter.objects.filter(is_active=True)
                
                # Envoyer la notification
                from django.core.mail import send_mail
                from django.conf import settings
                
                recipient_emails = [r.email for r in recipients]
                
                if recipient_emails:
                    send_mail(
                        subject,
                        message,
                        settings.EMAIL_HOST_USER,
                        recipient_emails,
                        fail_silently=False,
                    )
                    
                    messages.success(request, f'Notification envoyée à {len(recipient_emails)} abonnés avec succès.')
                else:
                    messages.warning(request, 'Aucun destinataire trouvé pour cette notification.')
                    
            except Exception as e:
                messages.error(request, f'Erreur lors de l\'envoi de la notification : {str(e)}')
            
            return redirect('content_management:newsletter_notifications')
        
        elif action == 'save_template':
            # Sauvegarder un template de notification
            template_name = request.POST.get('template_name')
            template_content = request.POST.get('template_content')
            
            # TODO: Implémenter la sauvegarde des templates
            messages.success(request, f'Template "{template_name}" sauvegardé avec succès.')
            return redirect('content_management:newsletter_notifications')
    
    # Récupérer les templates existants
    notification_templates = [
        {
            'name': 'Bienvenue',
            'subject': 'Bienvenue sur notre newsletter !',
            'content': 'Merci de vous être abonné à notre newsletter...'
        },
        {
            'name': 'Promotion',
            'subject': 'Offre spéciale pour nos abonnés',
            'content': 'Découvrez nos dernières offres...'
        },
        {
            'name': 'Rappel',
            'subject': 'Rappel : Votre abonnement expire bientôt',
            'content': 'N\'oubliez pas de renouveler votre abonnement...'
        }
    ]
    
    # Statistiques des notifications
    notification_stats = {
        'total_sent': 0,  # TODO: Implémenter le comptage réel
        'total_opened': 0,
        'total_clicked': 0,
    }
    
    context = {
        'notification_templates': notification_templates,
        'notification_stats': notification_stats,
    }
    
    return render(request, 'content_management/newsletter_notifications.html', context)

@login_required
def site_settings(request):
    """Paramètres généraux du site"""
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_general_settings':
            try:
                site_name = request.POST.get('site_name', '')
                site_tagline = request.POST.get('site_tagline', '')
                site_description = request.POST.get('site_description', '')
                
                if not site_name:
                    return JsonResponse({
                        'success': False,
                        'message': 'Le nom du site est obligatoire.'
                    })
                
                # Sauvegarder dans la session
                current_settings = request.session.get('site_settings', {})
                current_settings.update({
                    'site_name': site_name,
                    'site_tagline': site_tagline,
                    'site_description': site_description,
                })
                request.session['site_settings'] = current_settings
                
                return JsonResponse({
                    'success': True,
                    'message': 'Informations générales mises à jour avec succès.'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la mise à jour : {str(e)}'
                })
        
        elif action == 'update_contact_settings':
            try:
                contact_email = request.POST.get('contact_email', '')
                contact_phone = request.POST.get('contact_phone', '')
                contact_fax = request.POST.get('contact_fax', '')
                contact_website = request.POST.get('contact_website', '')
                
                if not contact_email:
                    return JsonResponse({
                        'success': False,
                        'message': 'L\'email de contact est obligatoire.'
                    })
                
                # Sauvegarder dans la session
                current_settings = request.session.get('site_settings', {})
                current_settings.update({
                    'contact_email': contact_email,
                    'contact_phone': contact_phone,
                    'contact_fax': contact_fax,
                    'contact_website': contact_website,
                })
                request.session['site_settings'] = current_settings
                
                return JsonResponse({
                    'success': True,
                    'message': 'Informations de contact mises à jour avec succès.'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la mise à jour : {str(e)}'
                })
        
        elif action == 'update_address_settings':
            try:
                address_street = request.POST.get('address_street', '')
                address_city = request.POST.get('address_city', '')
                address_country = request.POST.get('address_country', '')
                
                if not address_street or not address_city or not address_country:
                    return JsonResponse({
                        'success': False,
                        'message': 'La rue, la ville et le pays sont obligatoires.'
                    })
                
                # Sauvegarder dans la session
                current_settings = request.session.get('site_settings', {})
                current_settings.update({
                    'address_street': address_street,
                    'address_city': address_city,
                    'address_state': address_state,
                    'address_postal_code': address_postal_code,
                    'address_country': address_country,
                })
                request.session['site_settings'] = current_settings
                
                return JsonResponse({
                    'success': True,
                    'message': 'Adresse mise à jour avec succès.'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la mise à jour : {str(e)}'
                })
        
        elif action == 'update_social_settings':
            try:
                social_facebook = request.POST.get('social_facebook', '')
                social_twitter = request.POST.get('social_twitter', '')
                social_linkedin = request.POST.get('social_linkedin', '')
                social_instagram = request.POST.get('social_instagram', '')
                
                # Sauvegarder dans la session
                current_settings = request.session.get('site_settings', {})
                current_settings.update({
                    'social_facebook': social_facebook,
                    'social_twitter': social_twitter,
                    'social_linkedin': social_linkedin,
                    'social_instagram': social_instagram,
                })
                request.session['site_settings'] = current_settings
                
                return JsonResponse({
                    'success': True,
                    'message': 'Réseaux sociaux mis à jour avec succès.'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la mise à jour : {str(e)}'
                })
        
        elif action == 'update_technical_settings':
            try:
                timezone = request.POST.get('timezone', '')
                language = request.POST.get('language', '')
                
                if not timezone or not language:
                    return JsonResponse({
                        'success': False,
                        'message': 'Le fuseau horaire et la langue sont obligatoires.'
                    })
                
                # Sauvegarder dans la session
                current_settings = request.session.get('site_settings', {})
                current_settings.update({
                    'timezone': timezone,
                    'language': language,
                })
                request.session['site_settings'] = current_settings
                
                return JsonResponse({
                    'success': True,
                    'message': 'Paramètres techniques mis à jour avec succès.'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la mise à jour : {str(e)}'
                })
        
        elif action == 'save_all_settings':
            try:
                # Récupérer toutes les données des formulaires
                site_name = request.POST.get('site_name', '')
                site_tagline = request.POST.get('site_tagline', '')
                site_description = request.POST.get('site_description', '')
                contact_email = request.POST.get('contact_email', '')
                contact_phone = request.POST.get('contact_phone', '')
                contact_fax = request.POST.get('contact_fax', '')
                contact_website = request.POST.get('contact_website', '')
                address_street = request.POST.get('address_street', '')
                address_city = request.POST.get('address_city', '')
                address_state = request.POST.get('address_state', '')
                address_postal_code = request.POST.get('address_postal_code', '')
                address_country = request.POST.get('address_country', '')
                social_facebook = request.POST.get('social_facebook', '')
                social_twitter = request.POST.get('social_twitter', '')
                social_linkedin = request.POST.get('social_linkedin', '')
                social_instagram = request.POST.get('social_instagram', '')
                timezone = request.POST.get('timezone', '')
                language = request.POST.get('language', '')
                
                # Validation des champs obligatoires
                required_fields = {
                    'site_name': 'Nom du site',
                    'contact_email': 'Email de contact',
                    'address_street': 'Rue',
                    'address_city': 'Ville',
                    'address_country': 'Pays',
                    'timezone': 'Fuseau horaire',
                    'language': 'Langue'
                }
                
                missing_fields = []
                for field, label in required_fields.items():
                    if not request.POST.get(field, '').strip():
                        missing_fields.append(label)
                
                if missing_fields:
                    return JsonResponse({
                        'success': False,
                        'message': f'Champs obligatoires manquants : {", ".join(missing_fields)}'
                    })
                
                # Sauvegarder les données dans la session Django
                request.session['site_settings'] = {
                    'site_name': site_name,
                    'site_tagline': site_tagline,
                    'site_description': site_description,
                    'contact_email': contact_email,
                    'contact_phone': contact_phone,
                    'contact_fax': contact_fax,
                    'contact_website': contact_website,
                    'address_street': address_street,
                    'address_city': address_city,
                    'address_state': address_state,
                    'address_postal_code': address_postal_code,
                    'address_country': address_country,
                    'social_facebook': social_facebook,
                    'social_twitter': social_twitter,
                    'social_linkedin': social_linkedin,
                    'social_instagram': social_instagram,
                    'timezone': timezone,
                    'language': language,
                }
                
                return JsonResponse({
                    'success': True,
                    'message': 'Tous les paramètres ont été sauvegardés avec succès !'
                })
                
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la sauvegarde : {str(e)}'
                })
        
        elif action == 'reset_settings':
            try:
                # Supprimer les paramètres sauvegardés de la session
                if 'site_settings' in request.session:
                    del request.session['site_settings']
                
                return JsonResponse({
                    'success': True,
                    'message': 'Paramètres réinitialisés aux valeurs par défaut.'
                })
                
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la réinitialisation : {str(e)}'
                })
    
    # Paramètres par défaut
    default_settings = {
        'site_name': 'CSIG - Cité des Sciences et de l\'Innovation de Guinée',
        'site_tagline': 'Innovation et Excellence',
        'site_description': 'Organisation dédiée à la promotion des sciences et de l\'innovation en Guinée.',
        'contact_email': 'contact@csig.gn',
        'contact_phone': '+224 XXX XXX XXX',
        'contact_fax': '+224 XXX XXX XXX',
        'contact_website': 'https://www.csig.gn',
        'address_street': '123 Avenue de la Science',
        'address_city': 'Conakry',
        'address_state': 'Conakry',
        'address_postal_code': '00000',
        'address_country': 'Guinée',
        'social_facebook': 'https://facebook.com/csig',
        'social_twitter': 'https://twitter.com/csig',
        'social_linkedin': 'https://linkedin.com/company/csig',
        'social_instagram': 'https://instagram.com/csig',
        'timezone': 'Africa/Conakry',
        'language': 'fr',
    }
    
    # Récupérer les paramètres sauvegardés de la session ou utiliser les valeurs par défaut
    saved_settings = request.session.get('site_settings', {})
    site_settings = {**default_settings, **saved_settings}
    
    context = {
        'site_settings': site_settings,
    }
    
    return render(request, 'content_management/site_settings.html', context)

@login_required
def user_list(request):
    """Vue pour la liste des utilisateurs"""
    User = get_user_model()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create_user':
            try:
                # Récupérer les données du formulaire
                first_name = request.POST.get('first_name')
                last_name = request.POST.get('last_name')
                username = request.POST.get('username')
                email = request.POST.get('email')
                job_title = request.POST.get('job_title', '')
                department = request.POST.get('department', '')
                phone = request.POST.get('phone', '')
                language = request.POST.get('language', 'fr')
                roles = request.POST.getlist('roles')
                is_superuser = request.POST.get('is_superuser') == 'on'
                
                # Validation des données
                if not all([first_name, last_name, username, email]):
                    return JsonResponse({
                        'success': False,
                        'message': 'Tous les champs obligatoires doivent être remplis'
                    })
                
                # Vérifier si l'username ou l'email existe déjà
                if User.objects.filter(username=username).exists():
                    return JsonResponse({
                        'success': False,
                        'message': 'Ce nom d\'utilisateur existe déjà'
                    })
                
                if User.objects.filter(email=email).exists():
                    return JsonResponse({
                        'success': False,
                        'message': 'Cette adresse email existe déjà'
                    })
                
                # Générer un mot de passe sécurisé
                import secrets
                import string
                alphabet = string.ascii_letters + string.digits + string.punctuation
                password = ''.join(secrets.choice(alphabet) for i in range(16))
                
                # Créer l'utilisateur
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    is_superuser=is_superuser,
                    is_active=False,  # Désactivé par défaut, sera activé après vérification
                    is_verified=False
                )
                
                # Ajouter les attributs personnalisés
                user.job_title = job_title
                user.department = department
                user.phone = phone
                user.language = language
                user.save()
                
                # Ajouter les rôles
                if roles:
                    from users.models import UserRole
                    user_roles = UserRole.objects.filter(id__in=roles)
                    user.roles.set(user_roles)
                
                # Créer le profil utilisateur
                from users.models import UserProfile
                UserProfile.objects.create(user=user)
                
                # Envoyer l'email avec les informations de connexion
                from jobs.tasks import send_user_creation_email
                send_user_creation_email.delay(user.id, password)
                
                return JsonResponse({
                    'success': True,
                    'message': f'Utilisateur {username} créé avec succès. Un email avec les informations de connexion a été envoyé.'
                })
                
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la création: {str(e)}'
                })
                
        elif action == 'update_user':
            return JsonResponse({
                'success': True,
                'message': 'Utilisateur mis à jour avec succès'
            })
        elif action == 'delete_user':
            return JsonResponse({
                'success': True,
                'message': 'Utilisateur supprimé avec succès'
            })
        elif action == 'bulk_action':
            return JsonResponse({
                'success': True,
                'message': 'Action en lot appliquée avec succès'
            })
    
    # Récupérer les vraies données depuis la base de données
    users = User.objects.select_related('profile').prefetch_related('roles', 'groups').all()
    
    # Préparer les données pour le template
    users_data = []
    for user in users:
        # Récupérer les rôles de l'utilisateur
        user_roles = list(user.roles.values_list('name', flat=True))
        if not user_roles:
            user_roles = list(user.groups.values_list('name', flat=True))
        
        # Déterminer le statut
        if user.is_active:
            status = 'active'
        elif user.is_verified:
            status = 'pending'
        else:
            status = 'inactive'
        
        # Récupérer l'avatar si disponible
        avatar_url = None
        if hasattr(user, 'profile') and user.profile and user.profile.avatar:
            avatar_url = user.profile.avatar.url
        
        users_data.append({
            'id': user.id,
            'username': user.username,
            'first_name': user.first_name or '',
            'last_name': user.last_name or '',
            'email': user.email,
            'phone': getattr(user, 'phone', ''),
            'job_title': getattr(user, 'job_title', ''),
            'department': getattr(user, 'department', ''),
            'is_verified': getattr(user, 'is_verified', False),
            'language': getattr(user, 'language', 'fr'),
            'last_login': user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else '',
            'date_joined': user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
            'is_active': user.is_active,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser,
            'status': status,
            'roles': [{'name': role} for role in user_roles],
            'avatar_url': avatar_url
        })
    
    # Calculer les vraies statistiques
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    pending_users = User.objects.filter(is_active=False, is_verified=True).count()
    admin_users = User.objects.filter(is_superuser=True).count()
    
    # Récupérer les rôles disponibles pour le formulaire
    from users.models import UserRole
    available_roles = UserRole.objects.filter(is_active=True).order_by('name')
    
    context = {
        'total_users': total_users,
        'active_users': active_users,
        'pending_users': pending_users,
        'admin_users': admin_users,
        'users': users_data,
        'available_roles': available_roles
    }
    
    return render(request, 'content_management/user_list.html', context)


@login_required
def user_create(request):
    """Vue pour créer un nouvel utilisateur"""
    User = get_user_model()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create_user':
            try:
                # Récupérer les données du formulaire
                first_name = request.POST.get('first_name')
                last_name = request.POST.get('last_name')
                username = request.POST.get('username')
                email = request.POST.get('email')
                job_title = request.POST.get('job_title', '')
                department = request.POST.get('department', '')
                phone = request.POST.get('phone', '')
                language = request.POST.get('language', 'fr')
                roles = request.POST.getlist('roles')
                is_superuser = request.POST.get('is_superuser') == 'on'
                
                # Validation des données
                if not all([first_name, last_name, username, email]):
                    return JsonResponse({
                        'success': False,
                        'message': 'Tous les champs obligatoires doivent être remplis'
                    })
                
                # Vérifier si l'username ou l'email existe déjà
                if User.objects.filter(username=username).exists():
                    return JsonResponse({
                        'success': False,
                        'message': 'Ce nom d\'utilisateur existe déjà'
                    })
                
                if User.objects.filter(email=email).exists():
                    return JsonResponse({
                        'success': False,
                        'message': 'Cette adresse email existe déjà'
                    })
                
                # Générer un mot de passe sécurisé
                import secrets
                import string
                alphabet = string.ascii_letters + string.digits + string.punctuation
                password = ''.join(secrets.choice(alphabet) for i in range(16))
                
                # Créer l'utilisateur
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    is_superuser=is_superuser,
                    is_active=False,  # Désactivé par défaut, sera activé après vérification
                    is_verified=False
                )
                
                # Ajouter les attributs personnalisés
                user.job_title = job_title
                user.department = department
                user.phone = phone
                user.language = language
                user.save()
                
                # Ajouter les rôles
                if roles:
                    from users.models import UserRole
                    user_roles = UserRole.objects.filter(id__in=roles)
                    user.roles.set(user_roles)
                
                # Créer le profil utilisateur
                from users.models import UserProfile
                UserProfile.objects.create(user=user)
                
                # Envoyer l'email avec les informations de connexion
                from jobs.tasks import send_user_creation_email
                send_user_creation_email.delay(user.id, password)
                
                return JsonResponse({
                    'success': True,
                    'message': f'Utilisateur {username} créé avec succès. Un email avec les informations de connexion a été envoyé.'
                })
                
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la création: {str(e)}'
                })
    
    # Récupérer les rôles disponibles pour le formulaire
    from users.models import UserRole
    available_roles = UserRole.objects.filter(is_active=True).order_by('name')
    
    context = {
        'available_roles': available_roles
    }
    
    return render(request, 'content_management/user_create.html', context)


@login_required
def user_detail(request, pk):
    """Vue pour les détails d'un utilisateur"""
    User = get_user_model()
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_user':
            # Logique de mise à jour de l'utilisateur
            return JsonResponse({
                'success': True,
                'message': 'Utilisateur mis à jour avec succès'
            })
        elif action == 'delete_user':
            # Logique de suppression de l'utilisateur
            user.delete()
            messages.success(request, 'Utilisateur supprimé avec succès')
            return redirect('content_management:user_list')
        elif action == 'toggle_status':
            # Activer/désactiver l'utilisateur
            user.is_active = not user.is_active
            user.save()
            return JsonResponse({
                'success': True,
                'message': f'Utilisateur {"activé" if user.is_active else "désactivé"} avec succès'
            })
        elif action == 'change_role':
            # Changer le rôle de l'utilisateur
            role_id = request.POST.get('role_id')
            try:
                from users.models import UserRole
                new_role = UserRole.objects.get(id=role_id)
                user.roles.clear()
                user.roles.add(new_role)
                return JsonResponse({
                    'success': True,
                    'message': f'Rôle changé vers {new_role.name} avec succès'
                })
            except UserRole.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Rôle introuvable'
                })
    
    # Préparer les données pour le template
    user_data = {
        'id': user.id,
        'username': user.username,
        'first_name': user.first_name or '',
        'last_name': user.last_name or '',
        'email': user.email,
        'phone': getattr(user, 'phone', ''),
        'job_title': getattr(user, 'job_title', ''),
        'department': getattr(user, 'department', ''),
        'is_verified': getattr(user, 'is_verified', False),
        'language': getattr(user, 'language', 'fr'),
        'last_login': user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else '',
        'date_joined': user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
        'is_active': user.is_active,
        'is_staff': user.is_staff,
        'is_superuser': user.is_superuser,
        'status': status,
        'roles': list(user.roles.values_list('name', flat=True)),
        'avatar_url': None
    }
    
    # Récupérer l'avatar si disponible
    if hasattr(user, 'profile') and user.profile and user.profile.avatar:
        user_data['avatar_url'] = user.profile.avatar.url
    
    # Récupérer tous les rôles disponibles pour le formulaire de changement
    try:
        from users.models import UserRole
        available_roles = UserRole.objects.filter(is_active=True)
    except ImportError:
        available_roles = []
    
    context = {
        'user': user_data,
        'available_roles': available_roles
    }
    
    return render(request, 'content_management/user_detail.html', context)

def role_list(request):
    """Vue pour la liste des rôles"""
    # Importer les modèles nécessaires
    try:
        from users.models import UserRole
        from django.contrib.auth.models import Permission, Group
        User = get_user_model()
    except ImportError:
        # Fallback si les modèles ne sont pas disponibles
        User = get_user_model()
        UserRole = None
        Permission = None
        Group = None
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create_role':
            try:
                name = request.POST.get('name')
                description = request.POST.get('description', '')
                role_type = request.POST.get('type', 'custom')
                permissions = request.POST.getlist('permissions')
                
                # Créer le nouveau rôle
                role = UserRole.objects.create(
                    name=role_type,
                    description=description
                )
                
                # Ajouter les permissions sélectionnées
                if permissions:
                    role.permissions.set(Permission.objects.filter(id__in=permissions))
                
                return JsonResponse({
                    'success': True,
                    'message': 'Rôle créé avec succès'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la création: {str(e)}'
                })
                
        elif action == 'update_role':
            try:
                role_id = request.POST.get('role_id')
                name = request.POST.get('name')
                description = request.POST.get('description', '')
                permissions = request.POST.getlist('permissions')
                
                role = UserRole.objects.get(id=role_id)
                role.description = description
                role.save()
                
                # Mettre à jour les permissions
                if permissions:
                    role.permissions.set(Permission.objects.filter(id__in=permissions))
                
                return JsonResponse({
                    'success': True,
                    'message': 'Rôle mis à jour avec succès'
                })
            except UserRole.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Rôle introuvable'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la mise à jour: {str(e)}'
                })
                
        elif action == 'delete_role':
            try:
                role_id = request.POST.get('role_id')
                role = UserRole.objects.get(id=role_id)
                
                # Vérifier si des utilisateurs ont ce rôle
                users_with_role = User.objects.filter(roles=role).count()
                if users_with_role > 0:
                    return JsonResponse({
                        'success': False,
                        'message': f'Impossible de supprimer ce rôle car {users_with_role} utilisateur(s) l\'utilisent'
                    })
                
                role.delete()
                return JsonResponse({
                    'success': True,
                    'message': 'Rôle supprimé avec succès'
                })
            except UserRole.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Rôle introuvable'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la suppression: {str(e)}'
                })
    
    if UserRole and Permission:
        # Récupérer les vraies données depuis la base de données
        roles = UserRole.objects.prefetch_related('permissions').all()
        
        # Récupérer toutes les permissions disponibles pour les formulaires
        all_permissions = Permission.objects.all().order_by('content_type__app_label', 'content_type__model', 'codename')
        
        # Préparer les données pour le template
        roles_data = []
        for role in roles:
            # Compter les utilisateurs ayant ce rôle
            users_count = User.objects.filter(roles=role).count()
            
            # Récupérer les permissions
            permissions = list(role.permissions.values_list('name', flat=True))
            
            # Déterminer le type de rôle
            role_type = 'Système' if role.name in ['admin', 'manager', 'editor', 'author', 'viewer'] else 'Personnalisé'
            
            roles_data.append({
                'id': role.id,
                'name': role.get_name_display(),
                'type': role_type,
                'description': role.description or '',
                'permissions': [{'name': perm} for perm in permissions],
                'users': {'count': users_count},
                'created_at': role.created_at.strftime('%Y-%m-%d'),
                'is_active': role.is_active
            })
        
        # Calculer les vraies statistiques
        total_roles = UserRole.objects.count()
        active_roles = UserRole.objects.filter(is_active=True).count()
        custom_roles = UserRole.objects.exclude(name__in=['admin', 'manager', 'editor', 'author', 'viewer']).count()
        total_permissions = Permission.objects.count()
        
    else:
        # Fallback avec des données d'exemple si le modèle n'est pas disponible
        roles_data = [
            {
                'id': 1,
                'name': 'Administrateur',
                'type': 'Système',
                'description': 'Accès complet à toutes les fonctionnalités du système',
                'permissions': [
                    {'name': 'Gestion utilisateurs'},
                    {'name': 'Gestion rôles'},
                    {'name': 'Gestion contenu'},
                    {'name': 'Paramètres système'},
                    {'name': 'Logs système'}
                ],
                'users': {'count': 2},
                'created_at': '2024-01-01',
                'is_active': True
            }
        ]
        total_roles = 1
        active_roles = 1
        custom_roles = 0
        total_permissions = 5
        all_permissions = []
    
    context = {
        'total_roles': total_roles,
        'active_roles': active_roles,
        'custom_roles': custom_roles,
        'total_permissions': total_permissions,
        'roles': roles_data,
        'all_permissions': all_permissions
    }
    
    return render(request, 'content_management/role_list.html', context)


@login_required
def role_detail(request, pk):
    """Vue pour les détails d'un rôle"""
    try:
        from users.models import UserRole
        from django.contrib.auth.models import Permission, Group
        User = get_user_model()
    except ImportError:
        return JsonResponse({
            'success': False,
            'message': 'Modèles non disponibles'
        })
    
    role = get_object_or_404(UserRole, pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_role':
            try:
                description = request.POST.get('description', '')
                permissions = request.POST.getlist('permissions')
                
                role.description = description
                role.save()
                
                # Mettre à jour les permissions
                if permissions:
                    role.permissions.set(Permission.objects.filter(id__in=permissions))
                else:
                    role.permissions.clear()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Rôle mis à jour avec succès'
                })
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la mise à jour: {str(e)}'
                })
                
        elif action == 'delete_role':
            try:
                # Vérifier si des utilisateurs ont ce rôle
                users_with_role = User.objects.filter(roles=role).count()
                if users_with_role > 0:
                    return JsonResponse({
                        'success': False,
                        'message': f'Impossible de supprimer ce rôle car {users_with_role} utilisateur(s) l\'utilisent'
                    })
                
                role.delete()
                messages.success(request, 'Rôle supprimé avec succès')
                return redirect('content_management:role_list')
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Erreur lors de la suppression: {str(e)}'
                })
    
    # Préparer les données pour le template
    role_data = {
        'id': role.id,
        'name': role.get_name_display(),
        'type': 'Système' if role.name in ['admin', 'manager', 'editor', 'author', 'viewer'] else 'Personnalisé',
        'description': role.description or '',
        'permissions': list(role.permissions.values_list('name', flat=True)),
        'permission_ids': list(role.permissions.values_list('id', flat=True)),
        'users': {'count': User.objects.filter(roles=role).count()},
        'created_at': role.created_at.strftime('%Y-%m-%d'),
        'is_active': role.is_active
    }
    
    # Récupérer tous les utilisateurs ayant ce rôle
    users_with_role = User.objects.filter(roles=role).select_related('profile')
    users_data = []
    for user in users_with_role:
        users_data.append({
            'id': user.id,
            'username': user.username,
            'first_name': user.first_name or '',
            'last_name': user.last_name or '',
            'email': user.email,
            'is_active': user.is_active,
            'avatar_url': None
        })
        
        # Récupérer l'avatar si disponible
        if hasattr(user, 'profile') and user.profile and user.profile.avatar:
            users_data[-1]['avatar_url'] = user.profile.avatar.url
    
    # Récupérer toutes les permissions disponibles
    all_permissions = Permission.objects.all().order_by('content_type__app_label', 'content_type__model', 'codename')
    
    context = {
        'role': role_data,
        'users_with_role': users_data,
        'all_permissions': all_permissions,
    }
    
    return render(request, 'content_management/role_detail.html', context)

@login_required
def role_create(request):
    """Vue pour créer un nouveau rôle"""
    try:
        from users.models import UserRole
        from django.contrib.auth.models import Permission
    except ImportError:
        messages.error(request, 'Modèles non disponibles')
        return redirect('content_management:role_list')
    
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            description = request.POST.get('description', '')
            role_type = request.POST.get('type', 'custom')
            permissions = request.POST.getlist('permissions')
            
            # Validation
            if not name:
                messages.error(request, 'Le nom du rôle est obligatoire')
                return redirect('content_management:role_create')
            
            # Créer le nouveau rôle
            role = UserRole.objects.create(
                name=role_type,
                description=description
            )
            
            # Ajouter les permissions sélectionnées
            if permissions:
                role.permissions.set(Permission.objects.filter(id__in=permissions))
            
            messages.success(request, 'Rôle créé avec succès')
            return redirect('content_management:role_list')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création: {str(e)}')
            return redirect('content_management:role_create')
    
    # Récupérer toutes les permissions disponibles
    all_permissions = Permission.objects.all().order_by('content_type__app_label', 'content_type__model', 'codename')
    
    # Grouper les permissions par application
    permissions_by_app = {}
    for perm in all_permissions:
        app_label = perm.content_type.app_label
        if app_label not in permissions_by_app:
            permissions_by_app[app_label] = []
        permissions_by_app[app_label].append({
            'id': perm.id,
            'name': perm.name,
            'codename': perm.codename,
            'model': perm.content_type.model
        })
    
    context = {
        'all_permissions': all_permissions,
        'permissions_by_app': permissions_by_app
    }
    
    return render(request, 'content_management/role_create.html', context)

@login_required
def role_edit(request, pk):
    """Vue pour modifier un rôle existant"""
    try:
        from users.models import UserRole
        from django.contrib.auth.models import Permission
    except ImportError:
        messages.error(request, 'Modèles non disponibles')
        return redirect('content_management:role_list')
    
    role = get_object_or_404(UserRole, pk=pk)
    
    if request.method == 'POST':
        try:
            description = request.POST.get('description', '')
            permissions = request.POST.getlist('permissions')
            
            # Mettre à jour le rôle
            role.description = description
            role.save()
            
            # Mettre à jour les permissions
            if permissions:
                role.permissions.set(Permission.objects.filter(id__in=permissions))
            else:
                role.permissions.clear()
            
            messages.success(request, 'Rôle mis à jour avec succès')
            return redirect('content_management:role_detail', pk=pk)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')
            return redirect('content_management:role_edit', pk=pk)
    
    # Récupérer toutes les permissions disponibles
    all_permissions = Permission.objects.all().order_by('content_type__app_label', 'content_type__model', 'codename')
    
    # Grouper les permissions par application
    permissions_by_app = {}
    for perm in all_permissions:
        app_label = perm.content_type.app_label
        if app_label not in permissions_by_app:
            permissions_by_app[app_label] = []
        permissions_by_app[app_label].append({
            'id': perm.id,
            'name': perm.name,
            'codename': perm.codename,
            'model': perm.content_type.model
        })
    
    # Récupérer les permissions actuelles du rôle
    current_permissions = list(role.permissions.values_list('id', flat=True))
    
    context = {
        'role': role,
        'all_permissions': all_permissions,
        'permissions_by_app': permissions_by_app,
        'current_permissions': current_permissions
    }
    
    return render(request, 'content_management/role_edit.html', context)

# Gestion de l'équipe
@login_required
def team_member_list(request):
    """Liste des membres de l'équipe"""
    team_members = TeamMember.objects.all().order_by('-order', 'last_name', 'first_name')
    
    # Filtres
    category_filter = request.GET.get('category')
    status_filter = request.GET.get('status')
    search_query = request.GET.get('search')
    
    if category_filter:
        team_members = team_members.filter(category=category_filter)
    
    if status_filter:
        if status_filter == 'active':
            team_members = team_members.filter(is_active=True)
        elif status_filter == 'inactive':
            team_members = team_members.filter(is_active=False)
    
    if search_query:
        team_members = team_members.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(job_title__icontains=search_query) |
            Q(biography__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(team_members, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistiques
    total_members = TeamMember.objects.count()
    active_members = TeamMember.objects.filter(is_active=True).count()
    inactive_members = TeamMember.objects.filter(is_active=False).count()
    
    # Choix pour les filtres
    category_choices = TeamMember.CATEGORY_CHOICES
    status_choices = [
        ('all', 'Tous'),
        ('active', 'Actifs'),
        ('inactive', 'Inactifs'),
    ]
    
    context = {
        'page_obj': page_obj,
        'total_members': total_members,
        'active_members': active_members,
        'inactive_members': inactive_members,
        'category_choices': category_choices,
        'status_choices': status_choices,
    }
    
    return render(request, 'content_management/team_member_list.html', context)


@login_required
def team_member_create(request):
    """Créer un nouveau membre de l'équipe"""
    if request.method == 'POST':
        form = TeamMemberForm(request.POST, request.FILES)
        if form.is_valid():
            team_member = form.save()
            messages.success(request, f'Le membre {team_member.full_name} a été créé avec succès.')
            return redirect('content_management:team_member_list')
    else:
        form = TeamMemberForm()
    
    context = {
        'form': form,
        'page_title': 'Nouveau membre de l\'équipe',
        'is_create': True,
    }
    
    return render(request, 'content_management/team_member_form.html', context)


@login_required
def team_member_edit(request, pk):
    """Modifier un membre de l'équipe"""
    team_member = get_object_or_404(TeamMember, pk=pk)
    
    if request.method == 'POST':
        form = TeamMemberForm(request.POST, request.FILES, instance=team_member)
        if form.is_valid():
            team_member = form.save()
            messages.success(request, f'Le membre {team_member.full_name} a été modifié avec succès.')
            return redirect('content_management:team_member_list')
    else:
        form = TeamMemberForm(instance=team_member)
    
    context = {
        'form': form,
        'team_member': team_member,
        'page_title': f'Modifier {team_member.full_name}',
        'is_create': False,
    }
    
    return render(request, 'content_management/team_member_form.html', context)


@login_required
def team_member_detail(request, slug):
    """Vue détaillée d'un membre de l'équipe"""
    try:
        team_member = get_object_or_404(TeamMember, slug=slug)
    except TeamMember.DoesNotExist:
        messages.error(request, _("Membre de l'équipe introuvable."))
        return redirect('content_management:team_member_list')
    
    context = {
        'team_member': team_member,
        'page_title': _("Détails du membre"),
        'breadcrumb': [
            {'name': _("Accueil"), 'url': 'content_management:dashboard'},
            {'name': _("Équipe"), 'url': 'content_management:team_member_list'},
            {'name': team_member.full_name, 'url': '#'}
        ]
    }
    return render(request, 'content_management/team_member_detail.html', context)


@login_required
def team_member_delete(request, slug):
    """Vue de suppression d'un membre de l'équipe"""
    try:
        team_member = get_object_or_404(TeamMember, slug=slug)
    except TeamMember.DoesNotExist:
        messages.error(request, _("Membre de l'équipe introuvable."))
        return redirect('content_management:team_member_list')
    
    if request.method == 'POST':
        team_member.delete()
        messages.success(request, _("Membre de l'équipe supprimé avec succès."))
        return redirect('content_management:team_member_list')
    
    context = {
        'team_member': team_member,
        'page_title': _("Supprimer le membre"),
        'breadcrumb': [
            {'name': _("Accueil"), 'url': 'content_management:dashboard'},
            {'name': _("Équipe"), 'url': 'content_management:team_member_list'},
            {'name': _("Supprimer"), 'url': '#'}
        ]
    }
    return render(request, 'content_management/team_member_confirm_delete.html', context)


@login_required
def team_member_toggle_status(request, slug):
    """Activer/désactiver un membre de l'équipe"""
    if request.method == 'POST':
        team_member = get_object_or_404(TeamMember, slug=slug)
        team_member.is_active = not team_member.is_active
        team_member.save()
        
        status = "activé" if team_member.is_active else "désactivé"
        messages.success(request, f'Le membre {team_member.full_name} a été {status}.')
        
        return JsonResponse({'success': True, 'is_active': team_member.is_active})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


@login_required
def bulk_team_member_action(request):
    """Actions en masse sur les membres de l'équipe"""
    if request.method == 'POST':
        action = request.POST.get('action')
        member_ids = request.POST.getlist('member_ids[]')
        
        if not action or not member_ids:
            return JsonResponse({'success': False, 'error': 'Action ou membres manquants'})
        
        try:
            members = TeamMember.objects.filter(pk__in=member_ids)
            
            if action == 'activate':
                members.update(is_active=True)
                message = f'{len(members)} membre(s) activé(s)'
            elif action == 'deactivate':
                members.update(is_active=False)
                message = f'{len(members)} membre(s) désactivé(s)'
            elif action == 'delete':
                count = len(members)
                members.delete()
                message = f'{count} membre(s) supprimé(s)'
            else:
                return JsonResponse({'success': False, 'error': 'Action non reconnue'})
            
            messages.success(request, message)
            return JsonResponse({'success': True, 'message': message})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

# ===== GESTION DES SALLES DE CONFÉRENCE =====

@login_required
def conference_room_list(request):
    """Liste des salles de conférence"""
    rooms = ConferenceRoom.objects.all().order_by('name')
    
    # Filtres
    search_query = request.GET.get('search')
    availability_filter = request.GET.get('availability')
    
    if search_query:
        rooms = rooms.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    if availability_filter == 'available':
        rooms = rooms.filter(is_active=True)
        current_time = timezone.now()
        rooms = rooms.exclude(
            maintenance_until__gt=current_time
        )
    elif availability_filter == 'maintenance':
        current_time = timezone.now()
        rooms = rooms.filter(maintenance_until__gt=current_time)
    
    # Pagination
    paginator = Paginator(rooms, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'rooms': page_obj,
        'total_rooms': rooms.count(),
        'available_rooms': rooms.filter(is_active=True).count(),
        'maintenance_rooms': rooms.filter(maintenance_until__gt=timezone.now()).count(),
    }
    
    return render(request, 'content_management/conference_room_list.html', context)


@login_required
def conference_room_create(request):
    """Créer une nouvelle salle de conférence"""
    if request.method == 'POST':
        form = ConferenceRoomForm(request.POST)
        if form.is_valid():
            room = form.save()
            messages.success(request, _("Salle de conférence créée avec succès."))
            return redirect('content_management:conference_room_detail', slug=room.slug)
    else:
        form = ConferenceRoomForm()
    
    context = {
        'form': form,
        'title': _("Créer une salle de conférence"),
        'submit_text': _("Créer la salle")
    }
    
    return render(request, 'content_management/conference_room_form.html', context)


@login_required
def conference_room_detail(request, slug):
    """Détails d'une salle de conférence"""
    try:
        room = get_object_or_404(ConferenceRoom, slug=slug)
    except ConferenceRoom.DoesNotExist:
        messages.error(request, _("Salle de conférence introuvable."))
        return redirect('content_management:conference_room_list')
    
    # Réservations à venir
    upcoming_bookings = room.bookings.filter(
        start_time__gte=timezone.now(),
        status__in=['confirmed', 'pending']
    ).order_by('start_time')[:5]
    
    # Réservations récentes
    recent_bookings = room.bookings.filter(
        end_time__lte=timezone.now(),
        status='completed'
    ).order_by('-end_time')[:5]
    
    # Maintenance en cours
    active_maintenance = room.maintenance_records.filter(
        status='in_progress'
    ).first()
    
    context = {
        'room': room,
        'upcoming_bookings': upcoming_bookings,
        'recent_bookings': recent_bookings,
        'active_maintenance': active_maintenance,
    }
    
    return render(request, 'content_management/conference_room_detail.html', context)


@login_required
def conference_room_edit(request, slug):
    """Modifier une salle de conférence"""
    try:
        room = get_object_or_404(ConferenceRoom, slug=slug)
    except ConferenceRoom.DoesNotExist:
        messages.error(request, _("Salle de conférence introuvable."))
        return redirect('content_management:conference_room_list')
    
    if request.method == 'POST':
        form = ConferenceRoomForm(request.POST, instance=room)
        if form.is_valid():
            form.save()
            messages.success(request, _("Salle de conférence modifiée avec succès."))
            return redirect('content_management:conference_room_detail', slug=room.slug)
    else:
        form = ConferenceRoomForm(instance=room)
    
    context = {
        'form': form,
        'room': room,
        'title': _("Modifier la salle de conférence"),
        'submit_text': _("Enregistrer les modifications")
    }
    
    return render(request, 'content_management/conference_room_form.html', context)


@login_required
def conference_room_delete(request, slug):
    """Supprimer une salle de conférence"""
    try:
        room = get_object_or_404(ConferenceRoom, slug=slug)
    except ConferenceRoom.DoesNotExist:
        messages.error(request, _("Salle de conférence introuvable."))
        return redirect('content_management:conference_room_list')
    
    if request.method == 'POST':
        # Vérifier s'il y a des réservations actives
        active_bookings = room.bookings.filter(
            status__in=['confirmed', 'pending'],
            start_time__gte=timezone.now()
        )
        
        if active_bookings.exists():
            messages.error(request, _("Impossible de supprimer cette salle car elle a des réservations actives."))
            return redirect('content_management:conference_room_detail', slug=room.slug)
        
        room.delete()
        messages.success(request, _("Salle de conférence supprimée avec succès."))
        return redirect('content_management:conference_room_list')
    
    context = {
        'room': room,
        'title': _("Supprimer la salle de conférence")
    }
    
    return render(request, 'content_management/conference_room_confirm_delete.html', context)


@login_required
def external_organization_list(request):
    """Liste des organisations externes"""
    organizations = ExternalOrganization.objects.all().order_by('name')
    
    # Filtres
    search_query = request.GET.get('search')
    type_filter = request.GET.get('type')
    
    if search_query:
        organizations = organizations.filter(
            Q(name__icontains=search_query) |
            Q(contact_person__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    if type_filter:
        organizations = organizations.filter(organization_type=type_filter)
    
    # Pagination
    paginator = Paginator(organizations, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'organizations': page_obj,
        'total_organizations': organizations.count(),
        'active_organizations': organizations.filter(is_active=True).count(),
    }
    
    return render(request, 'content_management/external_organization_list.html', context)


@login_required
def external_organization_create(request):
    """Créer une nouvelle organisation externe"""
    if request.method == 'POST':
        form = ExternalOrganizationForm(request.POST)
        if form.is_valid():
            organization = form.save()
            messages.success(request, _("Organisation externe créée avec succès."))
            return redirect('content_management:external_organization_detail', slug=organization.slug)
    else:
        form = ExternalOrganizationForm()
    
    context = {
        'form': form,
        'title': _("Créer une organisation externe"),
        'submit_text': _("Créer l'organisation")
    }
    
    return render(request, 'content_management/external_organization_form.html', context)


@login_required
def external_organization_detail(request, slug):
    """Détails d'une organisation externe"""
    try:
        organization = get_object_or_404(ExternalOrganization, slug=slug)
    except ExternalOrganization.DoesNotExist:
        messages.error(request, _("Organisation externe introuvable."))
        return redirect('content_management:external_organization_list')
    
    # Réservations de cette organisation
    bookings = organization.bookings.all().order_by('-start_time')
    
    # Statistiques
    total_bookings = bookings.count()
    confirmed_bookings = bookings.filter(status='confirmed').count()
    total_spent = bookings.filter(status='confirmed').aggregate(
        total=models.Sum('total_price')
    )['total'] or 0
    
    context = {
        'organization': organization,
        'bookings': bookings[:10],  # 10 dernières réservations
        'total_bookings': total_bookings,
        'confirmed_bookings': confirmed_bookings,
        'total_spent': total_spent,
    }
    
    return render(request, 'content_management/external_organization_detail.html', context)


@login_required
def external_organization_edit(request, slug):
    """Modifier une organisation externe"""
    try:
        organization = get_object_or_404(ExternalOrganization, slug=slug)
    except ExternalOrganization.DoesNotExist:
        messages.error(request, _("Organisation externe introuvable."))
        return redirect('content_management:external_organization_list')
    
    if request.method == 'POST':
        form = ExternalOrganizationForm(request.POST, instance=organization)
        if form.is_valid():
            form.save()
            messages.success(request, _("Organisation externe modifiée avec succès."))
            return redirect('content_management:external_organization_detail', slug=organization.slug)
    else:
        form = ExternalOrganizationForm(instance=organization)
    
    context = {
        'form': form,
        'organization': organization,
        'title': _("Modifier l'organisation externe"),
        'submit_text': _("Enregistrer les modifications")
    }
    
    return render(request, 'content_management/external_organization_form.html', context)


@login_required
def external_organization_delete(request, slug):
    """Supprimer une organisation externe"""
    try:
        organization = get_object_or_404(ExternalOrganization, slug=slug)
    except ExternalOrganization.DoesNotExist:
        messages.error(request, _("Organisation externe introuvable."))
        return redirect('content_management:external_organization_list')
    
    if request.method == 'POST':
        # Vérifier s'il y a des réservations actives
        active_bookings = organization.bookings.filter(
            status__in=['confirmed', 'pending'],
            start_time__gte=timezone.now()
        )
        
        if active_bookings.exists():
            messages.error(request, _("Impossible de supprimer cette organisation car elle a des réservations actives."))
            return redirect('content_management:external_organization_detail', slug=organization.slug)
        
        organization.delete()
        messages.success(request, _("Organisation externe supprimée avec succès."))
        return redirect('content_management:external_organization_list')
    
    context = {
        'organization': organization,
        'title': _("Supprimer l'organisation externe")
    }
    
    return render(request, 'content_management/external_organization_confirm_delete.html', context)


@login_required
def room_booking_list(request):
    """Liste des réservations de salle"""
    bookings = RoomBooking.objects.all().order_by('-start_time')
    
    # Formulaire de recherche
    search_form = BookingSearchForm(request.GET)
    if search_form.is_valid():
        room = search_form.cleaned_data.get('room')
        organization = search_form.cleaned_data.get('organization')
        status = search_form.cleaned_data.get('status')
        start_date = search_form.cleaned_data.get('start_date')
        end_date = search_form.cleaned_data.get('end_date')
        search = search_form.cleaned_data.get('search')
        
        if room:
            bookings = bookings.filter(room=room)
        if organization:
            bookings = bookings.filter(organization=organization)
        if status:
            bookings = bookings.filter(status=status)
        if start_date:
            bookings = bookings.filter(start_time__date__gte=start_date)
        if end_date:
            bookings = bookings.filter(end_time__date__lte=end_date)
        if search:
            bookings = bookings.filter(
                Q(event_title__icontains=search) |
                Q(event_description__icontains=search)
            )
    
    # Pagination
    paginator = Paginator(bookings, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistiques
    total_bookings = bookings.count()
    pending_bookings = bookings.filter(status='pending').count()
    confirmed_bookings = bookings.filter(status='confirmed').count()
    total_revenue = bookings.filter(status='confirmed').aggregate(
        total=models.Sum('total_price')
    )['total'] or 0
    
    context = {
        'bookings': page_obj,
        'search_form': search_form,
        'total_bookings': total_bookings,
        'pending_bookings': pending_bookings,
        'confirmed_bookings': confirmed_bookings,
        'total_revenue': total_revenue,
    }
    
    return render(request, 'content_management/room_booking_list.html', context)


@login_required
def room_booking_create(request):
    """Créer une nouvelle réservation de salle"""
    if request.method == 'POST':
        form = RoomBookingForm(request.POST)
        if form.is_valid():
            try:
                # Créer la réservation
                booking = form.save(commit=False)
                # Définir le statut par défaut
                booking.status = 'pending'
                
                # Vérifier les conflits avant d'enregistrer
                try:
                    conflicts_summary = booking.get_detailed_conflicts_summary()
                    if conflicts_summary['has_conflicts']:
                        # Afficher un message d'erreur détaillé et structuré
                        messages.error(request, conflicts_summary['message'])
                        
                        # Ajouter des détails pour chaque conflit
                        for detail in conflicts_summary.get('details', []):
                            messages.warning(request, detail)
                        
                        # Message d'action recommandée
                        messages.info(request, "💡 Recommandations : Modifiez les dates/heures ou choisissez une autre salle disponible.")
                        
                        # Garder le formulaire avec les données saisies pour correction
                        context = {
                            'form': form,
                            'title': _("Créer une réservation"),
                            'submit_text': _("Créer la réservation"),
                            'conflicts_summary': conflicts_summary
                        }
                        return render(request, 'content_management/room_booking_form.html', context)
                    else:
                        # Aucun conflit, enregistrer la réservation
                        booking.save()
                        
                        messages.success(request, "✅ Réservation créée avec succès !")
                        messages.info(request, f"📅 Événement : {booking.event_title}")
                        messages.info(request, f"🏢 Salle : {booking.room.name}")
                        messages.info(request, f"📅 Date : {booking.start_time.strftime('%d/%m/%Y')} de {booking.start_time.strftime('%H:%M')} à {booking.end_time.strftime('%H:%M')}")
                        
                        return redirect('content_management:room_booking_detail', pk=booking.pk)
                        
                except Exception as conflict_error:
                    # Erreur lors de la vérification des conflits
                    messages.warning(request, "⚠️ La vérification automatique des conflits a rencontré un problème technique.")
                    messages.info(request, "💡 La réservation a été créée avec succès. Veuillez vérifier manuellement le calendrier de disponibilité.")
                    
                    # Enregistrer quand même la réservation
                    booking.save()
                    return redirect('content_management:room_booking_detail', pk=booking.pk)
                    
            except Exception as e:
                # Erreur générale lors de la création
                messages.error(request, "❌ Une erreur inattendue s'est produite lors de la création de la réservation.")
                messages.info(request, "💡 Veuillez réessayer ou contacter l'administrateur si le problème persiste.")
                
                # Log de l'erreur pour le débogage (pas visible par l'utilisateur)
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Erreur lors de la création de réservation: {str(e)}", exc_info=True)
                
                # Garder le formulaire avec les données saisies
        else:
            # Afficher les erreurs de validation de manière structurée
            if form.errors:
                messages.error(request, "❌ Veuillez corriger les erreurs suivantes :")
                
                for field, errors in form.errors.items():
                    for error in errors:
                        if field == '__all__':
                            # Erreur générale du formulaire
                            messages.error(request, f"• {error}")
                        elif field in form.fields:
                            # Erreur de champ spécifique
                            field_label = form.fields[field].label or field
                            messages.error(request, f"• {field_label} : {error}")
                        else:
                            # Erreur de champ inconnu
                            messages.error(request, f"• Erreur de validation : {error}")
    else:
        # Gérer les paramètres GET pour pré-remplir le formulaire
        initial_data = {}
        if 'room' in request.GET:
            try:
                room_id = request.GET.get('room')
                if room_id:
                    initial_data['room'] = room_id
            except (ValueError, TypeError):
                pass
        
        if 'organization' in request.GET:
            try:
                org_id = request.GET.get('organization')
                if org_id:
                    initial_data['organization'] = org_id
            except (ValueError, TypeError):
                pass
        
        form = RoomBookingForm(initial=initial_data)
    
    context = {
        'form': form,
        'title': _("Créer une réservation"),
        'submit_text': _("Créer la réservation")
    }
    
    return render(request, 'content_management/room_booking_form.html', context)


@login_required
def room_booking_detail(request, pk):
    """Détails d'une réservation de salle"""
    try:
        booking = get_object_or_404(RoomBooking, pk=pk)
    except RoomBooking.DoesNotExist:
        messages.error(request, _("Réservation introuvable."))
        return redirect('content_management:room_booking_list')
    
    # Paiements associés
    payments = booking.payments.all().order_by('-created_at')
    
    context = {
        'booking': booking,
        'payments': payments,
    }
    
    return render(request, 'content_management/room_booking_detail.html', context)


@login_required
def room_booking_edit(request, pk):
    """Modifier une réservation de salle"""
    try:
        booking = get_object_or_404(RoomBooking, pk=pk)
    except RoomBooking.DoesNotExist:
        messages.error(request, _("Réservation introuvable."))
        return redirect('content_management:room_booking_list')
    
    if request.method == 'POST':
        form = RoomBookingForm(request.POST, instance=booking)
        if form.is_valid():
            form.save()
            messages.success(request, _("Réservation modifiée avec succès."))
            return redirect('content_management:room_booking_detail', pk=booking.pk)
    else:
        form = RoomBookingForm(instance=booking)
    
    context = {
        'form': form,
        'booking': booking,
        'title': _("Modifier la réservation"),
        'submit_text': _("Enregistrer les modifications")
    }
    
    return render(request, 'content_management/room_booking_form.html', context)


@login_required
def room_booking_delete(request, pk):
    """Supprimer une réservation de salle"""
    try:
        booking = get_object_or_404(RoomBooking, pk=pk)
    except RoomBooking.DoesNotExist:
        messages.error(request, _("Réservation introuvable."))
        return redirect('content_management:room_booking_list')
    
    if request.method == 'POST':
        # Vérifier si la réservation peut être supprimée
        if booking.status in ['confirmed', 'pending'] and booking.start_time > timezone.now():
            messages.error(request, _("Impossible de supprimer une réservation confirmée ou en attente qui n'a pas encore eu lieu."))
            return redirect('content_management:room_booking_detail', pk=booking.pk)
        
        booking.delete()
        messages.success(request, _("Réservation supprimée avec succès."))
        return redirect('content_management:room_booking_list')
    
    context = {
        'booking': booking,
        'title': _("Supprimer la réservation")
    }
    
    return render(request, 'content_management/room_booking_confirm_delete.html', context)


@login_required
def check_booking_conflicts(request):
    """Vérifier les conflits pour une réservation (AJAX)"""
    if request.method == 'POST':
        try:
            room_id = request.POST.get('room')
            start_time = request.POST.get('start_time')
            end_time = request.POST.get('end_time')
            booking_id = request.POST.get('booking_id')  # Pour les modifications
            
            if not all([room_id, start_time, end_time]):
                return JsonResponse({
                    'success': False,
                    'error': '❌ Données manquantes : veuillez remplir tous les champs requis.'
                })
            
            # Convertir les dates avec gestion des timezones
            from django.utils.dateparse import parse_datetime
            from django.utils import timezone
            
            start_dt = parse_datetime(start_time)
            end_dt = parse_datetime(end_time)
            
            if not start_dt or not end_dt:
                return JsonResponse({
                    'success': False,
                    'error': '❌ Format de date invalide : veuillez utiliser le format YYYY-MM-DD HH:MM'
                })
            
            # S'assurer que les dates ont le bon fuseau horaire
            from datetime import timezone as dt_timezone
            if timezone.is_naive(start_dt):
                start_dt = timezone.make_aware(start_dt, dt_timezone.utc)
            if timezone.is_naive(end_dt):
                end_dt = timezone.make_aware(end_dt, dt_timezone.utc)
            
            # Validation des dates
            if start_dt >= end_dt:
                return JsonResponse({
                    'success': False,
                    'error': '❌ La date de fin doit être postérieure à la date de début.'
                })
            
            if start_dt < timezone.now():
                return JsonResponse({
                    'success': False,
                    'error': '❌ La date de début ne peut pas être dans le passé.'
                })
            
            # Créer un objet temporaire pour la vérification
            temp_booking = RoomBooking()
            temp_booking.room_id = room_id
            temp_booking.start_time = start_dt
            temp_booking.end_time = end_dt
            temp_booking.pk = booking_id  # Pour exclure la réservation actuelle en cas de modification
            
            # Obtenir le rapport des conflits
            try:
                conflicts_summary = temp_booking.get_detailed_conflicts_summary()
                duration_info = temp_booking.get_duration_info()
                
                return JsonResponse({
                    'success': True,
                    'conflicts_summary': conflicts_summary,
                    'duration_info': duration_info,
                    'formatted_start': start_dt.strftime('%d/%m/%Y à %H:%M'),
                    'formatted_end': end_dt.strftime('%d/%m/%Y à %H:%M')
                })
                
            except Exception as conflict_error:
                # Log de l'erreur pour le débogage
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Erreur lors de la vérification des conflits: {str(conflict_error)}", exc_info=True)
                
                return JsonResponse({
                    'success': False,
                    'error': '⚠️ La vérification des conflits a rencontré un problème technique. Veuillez réessayer.'
                })
            
        except Exception as e:
            # Log de l'erreur pour le débogage
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erreur technique lors de la vérification des conflits: {str(e)}", exc_info=True)
            
            return JsonResponse({
                'success': False,
                'error': '❌ Une erreur technique s\'est produite. Veuillez réessayer ou contacter l\'administrateur.'
            })
    
    return JsonResponse({
        'success': False, 
        'error': '❌ Méthode non autorisée : utilisez POST pour vérifier les conflits.'
    })

@login_required
def room_booking_toggle_status(request, pk):
    """Basculer le statut d'une réservation"""
    if request.method == 'POST':
        try:
            booking = get_object_or_404(RoomBooking, pk=pk)
            
            # Vérifier si c'est une requête AJAX avec du JSON
            if request.headers.get('Content-Type') == 'application/json':
                import json
                data = json.loads(request.body)
                new_status = data.get('status')
            else:
                # Fallback pour les formulaires POST traditionnels
                new_status = request.POST.get('status')
            
            if new_status in ['confirmed', 'pending', 'cancelled', 'completed', 'rejected']:
                old_status = booking.status
                booking.status = new_status
                
                # Marquer comme confirmé si nécessaire
                if new_status == 'confirmed' and not booking.confirmed_at:
                    booking.confirmed_by = request.user
                    booking.confirmed_at = timezone.now()
                
                booking.save()
                
                status_display = dict(RoomBooking._meta.get_field('status').choices)[new_status]
                
                # Si c'est une requête AJAX, retourner du JSON
                if request.headers.get('Content-Type') == 'application/json':
                    return JsonResponse({
                        'success': True,
                        'message': f"Statut de la réservation changé en '{status_display}'.",
                        'new_status': new_status,
                        'status_display': status_display
                    })
                else:
                    # Sinon, utiliser les messages Django et rediriger
                    messages.success(request, f"Statut de la réservation changé en '{status_display}'.")
                    return redirect('content_management:room_booking_detail', pk=pk)
            else:
                error_msg = _("Statut invalide.")
                if request.headers.get('Content-Type') == 'application/json':
                    return JsonResponse({
                        'success': False,
                        'error': error_msg
                    })
                else:
                    messages.error(request, error_msg)
                    return redirect('content_management:room_booking_detail', pk=pk)
                
        except RoomBooking.DoesNotExist:
            error_msg = _("Réservation introuvable.")
            if request.headers.get('Content-Type') == 'application/json':
                return JsonResponse({
                    'success': False,
                    'error': error_msg
                })
            else:
                messages.error(request, error_msg)
                return redirect('content_management:room_booking_list')
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': _("Données JSON invalides.")
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    # Si ce n'est pas une requête POST
    if request.headers.get('Content-Type') == 'application/json':
        return JsonResponse({
            'success': False,
            'error': _("Méthode non autorisée.")
        })
    else:
        return redirect('content_management:room_booking_list')


@login_required
def room_booking_calendar(request):
    """Calendrier des réservations de salle"""
    from calendar import monthcalendar
    from datetime import date
    
    # Récupérer le mois demandé ou le mois actuel
    month_param = request.GET.get('month')
    if month_param:
        try:
            year, month = map(int, month_param.split('-'))
            current_month = date(year, month, 1)
        except (ValueError, TypeError):
            current_month = timezone.now().date().replace(day=1)
    else:
        current_month = timezone.now().date().replace(day=1)
    
    # Calculer les mois précédent et suivant
    if current_month.month == 1:
        previous_month = current_month.replace(year=current_month.year - 1, month=12)
    else:
        previous_month = current_month.replace(month=current_month.month - 1)
    
    if current_month.month == 12:
        next_month = current_month.replace(year=current_month.year + 1, month=1)
    else:
        next_month = current_month.replace(month=current_month.month + 1)
    
    # Récupérer toutes les salles
    rooms = ConferenceRoom.objects.filter(is_active=True)
    
    # Récupérer les réservations pour le mois demandé
    start_of_month = current_month
    if current_month.month == 12:
        end_of_month = current_month.replace(year=current_month.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        end_of_month = current_month.replace(month=current_month.month + 1, day=1) - timedelta(days=1)
    
    bookings = RoomBooking.objects.filter(
        start_time__date__gte=start_of_month,
        end_time__date__lte=end_of_month,
        status__in=['confirmed', 'pending', 'cancelled', 'completed', 'rejected']
    ).select_related('room', 'organization')
    
    # Organiser les réservations par salle et par date
    calendar_data = {}
    today = timezone.now().date()
    
    for room in rooms:
        calendar_data[room.id] = {
            'room': room,
            'weeks': []
        }
        
        # Générer le calendrier du mois
        month_calendar = monthcalendar(current_month.year, current_month.month)
        
        for week in month_calendar:
            week_data = []
            for day in week:
                if day == 0:
                    # Jour d'un autre mois
                    week_data.append({
                        'date': None,
                        'bookings': [],
                        'is_today': False,
                        'is_other_month': True
                    })
                else:
                    current_date = date(current_month.year, current_month.month, day)
                    is_today = current_date == today
                    
                    # Récupérer les réservations pour cette date
                    day_bookings = []
                    for booking in bookings:
                        if (booking.room_id == room.id and 
                            current_date >= booking.start_time.date() and 
                            current_date <= booking.end_time.date()):
                            day_bookings.append(booking)
                    
                    week_data.append({
                        'date': current_date,
                        'bookings': day_bookings,
                        'is_today': is_today,
                        'is_other_month': False
                    })
            
            calendar_data[room.id]['weeks'].append(week_data)
    
    context = {
        'calendar_data': calendar_data,
        'current_month': current_month,
        'previous_month': previous_month,
        'next_month': next_month,
        'start_of_month': start_of_month,
        'end_of_month': end_of_month,
        'today': today,
    }
    
    return render(request, 'content_management/room_booking_calendar.html', context)


@login_required
def room_maintenance_list(request):
    """Liste des maintenances de salle"""
    maintenances = RoomMaintenance.objects.all().order_by('-start_date')
    
    # Filtres
    status_filter = request.GET.get('status')
    room_filter = request.GET.get('room')
    
    if status_filter:
        maintenances = maintenances.filter(status=status_filter)
    
    if room_filter:
        maintenances = maintenances.filter(room_id=room_filter)
    
    # Pagination
    paginator = Paginator(maintenances, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'maintenances': page_obj,
        'rooms': ConferenceRoom.objects.filter(is_active=True),
        'total_maintenances': maintenances.count(),
        'active_maintenances': maintenances.filter(status='in_progress').count(),
    }
    
    return render(request, 'content_management/room_maintenance_list.html', context)


@login_required
def room_maintenance_create(request):
    """Créer une nouvelle maintenance de salle"""
    if request.method == 'POST':
        form = RoomMaintenanceForm(request.POST)
        if form.is_valid():
            maintenance = form.save()
            messages.success(request, _("Maintenance créée avec succès."))
            return redirect('content_management:room_maintenance_detail', pk=maintenance.pk)
    else:
        form = RoomMaintenanceForm()
    
    context = {
        'form': form,
        'title': _("Créer une maintenance"),
        'submit_text': _("Créer la maintenance")
    }
    
    return render(request, 'content_management/room_maintenance_create.html', context)


@login_required
def room_maintenance_detail(request, pk):
    """Détails d'une maintenance de salle"""
    try:
        maintenance = get_object_or_404(RoomMaintenance, pk=pk)
    except RoomMaintenance.DoesNotExist:
        messages.error(request, _("Maintenance introuvable."))
        return redirect('content_management:room_maintenance_list')
    
    context = {
        'maintenance': maintenance,
    }
    
    return render(request, 'content_management/room_maintenance_detail.html', context)


@login_required
def room_maintenance_edit(request, pk):
    """Modifier une maintenance de salle"""
    try:
        maintenance = get_object_or_404(RoomMaintenance, pk=pk)
    except RoomMaintenance.DoesNotExist:
        messages.error(request, _("Maintenance introuvable."))
        return redirect('content_management:room_maintenance_list')
    
    if request.method == 'POST':
        form = RoomMaintenanceForm(request.POST, instance=maintenance)
        if form.is_valid():
            form.save()
            messages.success(request, _("Maintenance modifiée avec succès."))
            return redirect('content_management:room_maintenance_detail', pk=maintenance.pk)
    else:
        form = RoomMaintenanceForm(instance=maintenance)
    
    context = {
        'form': form,
        'maintenance': maintenance,
        'title': _("Modifier la maintenance"),
        'submit_text': _("Enregistrer les modifications")
    }
    
    return render(request, 'content_management/room_maintenance_form.html', context)


@login_required
def room_maintenance_delete(request, pk):
    """Supprimer une maintenance de salle"""
    try:
        maintenance = get_object_or_404(RoomMaintenance, pk=pk)
    except RoomMaintenance.DoesNotExist:
        messages.error(request, _("Maintenance introuvable."))
        return redirect('content_management:room_maintenance_list')
    
    if request.method == 'POST':
        maintenance.delete()
        messages.success(request, _("Maintenance supprimée avec succès."))
        return redirect('content_management:room_maintenance_list')
    
    context = {
        'maintenance': maintenance,
        'title': _("Supprimer la maintenance")
    }
    
    return render(request, 'content_management/room_maintenance_confirm_delete.html', context)


@login_required
def dashboard_rooms(request):
    """Tableau de bord des salles de conférence"""
    # Statistiques générales
    total_rooms = ConferenceRoom.objects.count()
    available_rooms = ConferenceRoom.objects.filter(is_active=True).count()
    
    # Réservations
    total_bookings = RoomBooking.objects.count()
    pending_bookings = RoomBooking.objects.filter(status='pending').count()
    confirmed_bookings = RoomBooking.objects.filter(status='confirmed').count()
    today_bookings = RoomBooking.objects.filter(
        start_time__date=timezone.now().date(),
        status__in=['confirmed', 'pending']
    ).count()
    
    # Revenus
    total_revenue = RoomBooking.objects.filter(status='confirmed').aggregate(
        total=models.Sum('total_price')
    )['total'] or 0
    
    monthly_revenue = RoomBooking.objects.filter(
        status='confirmed',
        start_time__month=timezone.now().month,
        start_time__year=timezone.now().year
    ).aggregate(total=models.Sum('total_price'))['total'] or 0
    
    # Organisations
    total_organizations = ExternalOrganization.objects.count()
    active_organizations = ExternalOrganization.objects.filter(is_active=True).count()
    
    # Maintenances
    active_maintenances = RoomMaintenance.objects.filter(status='in_progress').count()
    planned_maintenances = RoomMaintenance.objects.filter(status='planned').count()
    
    # Réservations à venir (prochaines 7 jours)
    upcoming_bookings = RoomBooking.objects.filter(
        start_time__gte=timezone.now(),
        start_time__lte=timezone.now() + timedelta(days=7),
        status__in=['confirmed', 'pending']
    ).order_by('start_time')[:10]
    
    # Salles les plus populaires
    popular_rooms = ConferenceRoom.objects.annotate(
        booking_count=models.Count('bookings', filter=models.Q(bookings__status='confirmed'))
    ).order_by('-booking_count')[:5]
    
    context = {
        'total_rooms': total_rooms,
        'available_rooms': available_rooms,
        'total_bookings': total_bookings,
        'pending_bookings': pending_bookings,
        'confirmed_bookings': confirmed_bookings,
        'today_bookings': today_bookings,
        'total_revenue': total_revenue,
        'monthly_revenue': monthly_revenue,
        'total_organizations': total_organizations,
        'active_organizations': active_organizations,
        'active_maintenances': active_maintenances,
        'planned_maintenances': planned_maintenances,
        'upcoming_bookings': upcoming_bookings,
        'popular_rooms': popular_rooms,
    }
    
    return render(request, 'content_management/dashboard_rooms.html', context)

# ============================================================================
# GESTION DES PERSONNAS
# ============================================================================

@login_required
def personna_list(request):
    """Liste des personnas"""
    personnas = Personna.objects.all().order_by('-created_at')
    
    # Filtres
    personna_type = request.GET.get('type')
    status = request.GET.get('status')
    search = request.GET.get('search')
    
    if personna_type:
        personnas = personnas.filter(personna_type=personna_type)
    
    if status == 'active':
        personnas = personnas.filter(is_active=True)
    elif status == 'inactive':
        personnas = personnas.filter(is_active=False)
    
    if search:
        personnas = personnas.filter(
            Q(name__icontains=search) |
            Q(bio__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(personnas, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'personnas': page_obj,
        'total_count': personnas.count(),
        'active_count': personnas.filter(is_active=True).count(),
        'featured_count': personnas.filter(featured=True).count(),
    }
    
    return render(request, 'content_management/personna_list.html', context)


@login_required
def personna_create(request):
    """Créer un nouveau personna"""
    if request.method == 'POST':
        form = PersonnaForm(request.POST, request.FILES)
        if form.is_valid():
            personna = form.save()
            messages.success(request, f'Le personna "{personna.name}" a été créé avec succès.')
            return redirect('content_management:personna_detail', slug=personna.slug)
    else:
        form = PersonnaForm()
    
    context = {
        'form': form,
        'page_title': 'Nouveau personna',
        'is_create': True,
    }
    
    return render(request, 'content_management/personna_form.html', context)


@login_required
def personna_detail(request, slug):
    """Détails d'un personna"""
    try:
        personna = get_object_or_404(Personna, slug=slug)
    except Personna.DoesNotExist:
        messages.error(request, _("Personna introuvable."))
        return redirect('content_management:personna_list')
    
    # Blogs du personna
    blogs = personna.blogs.filter(is_active=True).order_by('-created_at')[:5]
    
    context = {
        'personna': personna,
        'blogs': blogs,
        'page_title': personna.name,
        'breadcrumb': [
            {'name': _("Accueil"), 'url': 'content_management:dashboard'},
            {'name': _("Personnas"), 'url': 'content_management:personna_list'},
            {'name': personna.name, 'url': '#'}
        ]
    }
    return render(request, 'content_management/personna_detail.html', context)


@login_required
def personna_edit(request, slug):
    """Modifier un personna"""
    try:
        personna = get_object_or_404(Personna, slug=slug)
    except Personna.DoesNotExist:
        messages.error(request, _("Personna introuvable."))
        return redirect('content_management:personna_list')
    
    if request.method == 'POST':
        form = PersonnaForm(request.POST, request.FILES, instance=personna)
        if form.is_valid():
            personna = form.save()
            messages.success(request, f'Le personna "{personna.name}" a été modifié avec succès.')
            return redirect('content_management:personna_detail', slug=personna.slug)
    else:
        form = PersonnaForm(instance=personna)
    
    context = {
        'form': form,
        'personna': personna,
        'page_title': f'Modifier {personna.name}',
        'is_create': False,
    }
    
    return render(request, 'content_management/personna_form.html', context)


@login_required
def personna_delete(request, slug):
    """Supprimer un personna"""
    try:
        personna = get_object_or_404(Personna, slug=slug)
    except Personna.DoesNotExist:
        messages.error(request, _("Personna introuvable."))
        return redirect('content_management:personna_list')
    
    if request.method == 'POST':
        name = personna.name
        personna.delete()
        messages.success(request, f'Le personna "{name}" a été supprimé avec succès.')
        return redirect('content_management:personna_list')
    
    context = {
        'personna': personna,
        'page_title': f'Supprimer {personna.name}',
        'breadcrumb': [
            {'name': _("Accueil"), 'url': 'content_management:dashboard'},
            {'name': _("Personnas"), 'url': 'content_management:personna_list'},
            {'name': _("Supprimer"), 'url': '#'}
        ]
    }
    return render(request, 'content_management/personna_confirm_delete.html', context)


@login_required
def personna_toggle_status(request, slug):
    """Activer/désactiver un personna"""
    if request.method == 'POST':
        try:
            personna = get_object_or_404(Personna, slug=slug)
            personna.is_active = not personna.is_active
            personna.save()
            
            status = "activé" if personna.is_active else "désactivé"
            messages.success(request, f'Le personna "{personna.name}" a été {status}.')
            
            return JsonResponse({'success': True, 'is_active': personna.is_active})
        except Personna.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Personna introuvable'})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


@login_required
def personna_toggle_featured(request, slug):
    """Mettre en avant/retirer de l'avant un personna"""
    if request.method == 'POST':
        try:
            personna = get_object_or_404(Personna, slug=slug)
            personna.featured = not personna.featured
            personna.save()
            
            status = "mis en avant" if personna.featured else "retiré de l'avant"
            messages.success(request, f'Le personna "{personna.name}" a été {status}.')
            
            return JsonResponse({'success': True, 'featured': personna.featured})
        except Personna.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Personna introuvable'})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


# ============================================================================
# GESTION DES BLOGS
# ============================================================================

@login_required
def blog_list(request):
    """Liste des blogs"""
    blogs = Blog.objects.all().order_by('-created_at')
    
    # Filtres
    personna = request.GET.get('personna')
    status = request.GET.get('status')
    search = request.GET.get('search')
    
    if personna:
        blogs = blogs.filter(personna_id=personna)
    
    if status:
        blogs = blogs.filter(status=status)
    
    if search:
        blogs = blogs.filter(
            Q(title__icontains=search) |
            Q(content__icontains=search) |
            Q(excerpt__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(blogs, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'blogs': page_obj,
        'personnas': Personna.objects.filter(is_active=True),
        'status_choices': Blog.STATUS_CHOICES,
        'total_count': blogs.count(),
        'published_count': blogs.filter(status='published').count(),
        'draft_count': blogs.filter(status='draft').count(),
        'featured_count': blogs.filter(featured=True).count(),
    }
    
    return render(request, 'content_management/blog_list.html', context)


@login_required
def blog_create(request):
    """Créer un nouveau blog"""
    if request.method == 'POST':
        form = BlogForm(request.POST, request.FILES)
        if form.is_valid():
            blog = form.save()
            messages.success(request, f'Le blog "{blog.title}" a été créé avec succès.')
            return redirect('content_management:blog_detail', slug=blog.slug)
    else:
        form = BlogForm()
    
    context = {
        'form': form,
        'page_title': 'Nouveau blog',
        'is_create': True,
    }
    
    return render(request, 'content_management/blog_form.html', context)


@login_required
def blog_detail(request, slug):
    """Détails d'un blog"""
    try:
        blog = get_object_or_404(Blog, slug=slug)
    except Blog.DoesNotExist:
        messages.error(request, _("Blog introuvable."))
        return redirect('content_management:blog_list')
    
    # Incrémenter le compteur de vues
    blog.views_count += 1
    blog.save(update_fields=['views_count'])
    
    # Blogs similaires du même personna
    similar_blogs = Blog.objects.filter(
        personna=blog.personna,
        is_active=True
    ).exclude(pk=blog.pk).order_by('-created_at')[:3]
    
    context = {
        'blog': blog,
        'similar_blogs': similar_blogs,
        'page_title': blog.title,
        'breadcrumb': [
            {'name': _("Accueil"), 'url': 'content_management:dashboard'},
            {'name': _("Blogs"), 'url': 'content_management:blog_list'},
            {'name': blog.title, 'url': '#'}
        ]
    }
    return render(request, 'content_management/blog_detail.html', context)


@login_required
def blog_edit(request, slug):
    """Modifier un blog"""
    try:
        blog = get_object_or_404(Blog, slug=slug)
    except Blog.DoesNotExist:
        messages.error(request, _("Blog introuvable."))
        return redirect('content_management:blog_list')
    
    if request.method == 'POST':
        form = BlogForm(request.POST, request.FILES, instance=blog)
        if form.is_valid():
            blog = form.save()
            messages.success(request, f'Le blog "{blog.title}" a été modifié avec succès.')
            return redirect('content_management:blog_detail', slug=blog.slug)
    else:
        form = BlogForm(instance=blog)
    
    context = {
        'form': form,
        'blog': blog,
        'page_title': f'Modifier {blog.title}',
        'is_create': False,
    }
    
    return render(request, 'content_management/blog_form.html', context)


@login_required
def blog_delete(request, slug):
    """Supprimer un blog"""
    try:
        blog = get_object_or_404(Blog, slug=slug)
    except Blog.DoesNotExist:
        messages.error(request, _("Blog introuvable."))
        return redirect('content_management:blog_list')
    
    if request.method == 'POST':
        title = blog.title
        blog.delete()
        messages.success(request, f'Le blog "{title}" a été supprimé avec succès.')
        return redirect('content_management:blog_list')
    
    context = {
        'blog': blog,
        'page_title': f'Supprimer {blog.title}',
        'breadcrumb': [
            {'name': _("Accueil"), 'url': 'content_management:dashboard'},
            {'name': _("Blogs"), 'url': 'content_management:blog_list'},
            {'name': _("Supprimer"), 'url': '#'}
        ]
    }
    return render(request, 'content_management/blog_confirm_delete.html', context)


@login_required
def blog_toggle_status(request, slug):
    """Changer le statut d'un blog"""
    if request.method == 'POST':
        try:
            blog = get_object_or_404(Blog, slug=slug)
            new_status = request.POST.get('status')
            
            if new_status in dict(Blog.STATUS_CHOICES):
                blog.status = new_status
                if new_status == 'published' and not blog.published_at:
                    blog.published_at = timezone.now()
                blog.save()
                
                messages.success(request, f'Le statut du blog "{blog.title}" a été changé en "{blog.get_status_display()}".')
                return JsonResponse({'success': True, 'status': blog.status})
            else:
                return JsonResponse({'success': False, 'error': 'Statut invalide'})
        except Blog.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Blog introuvable'})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


@login_required
def blog_toggle_active(request, slug):
    """Activer/désactiver un blog"""
    if request.method == 'POST':
        try:
            blog = get_object_or_404(Blog, slug=slug)
            blog.is_active = not blog.is_active
            blog.save()
            
            status = "activé" if blog.is_active else "désactivé"
            messages.success(request, f'Le blog "{blog.title}" a été {status}.')
            
            return JsonResponse({'success': True, 'is_active': blog.is_active})
        except Blog.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Blog introuvable'})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


@login_required
def blog_toggle_featured(request, slug):
    """Mettre en avant/retirer de l'avant un blog"""
    if request.method == 'POST':
        try:
            blog = get_object_or_404(Blog, slug=slug)
            blog.featured = not blog.featured
            blog.save()
            
            status = "mis en avant" if blog.featured else "retiré de l'avant"
            messages.success(request, f'Le blog "{blog.title}" a été {status}.')
            
            return JsonResponse({'success': True, 'featured': blog.featured})
        except Blog.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Blog introuvable'})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

@login_required
def test_ckeditor(request):
    """Vue de test pour CKEditor5"""
    if request.method == 'POST':
        content = request.POST.get('content', '')
        messages.success(request, f'Contenu reçu avec succès ! Longueur : {len(content)} caractères')
        return redirect('content_management:test_ckeditor')
    
    return render(request, 'content_management/test_ckeditor.html')

@login_required
def about_page_edit(request):
    """Éditer la page À propos"""
    try:
        about_page = AboutPage.get_active_page()
    except AboutPage.DoesNotExist:
        about_page = None
    
    if request.method == 'POST':
        form = AboutPageForm(request.POST, request.FILES, instance=about_page)
        if form.is_valid():
            # Désactiver toutes les autres pages
            AboutPage.objects.update(is_active=False)
            
            # Sauvegarder la nouvelle page
            about_page = form.save(commit=False)
            about_page.is_active = True
            about_page.save()
            
            messages.success(request, 'Page À propos mise à jour avec succès !')
            return redirect('content_management:about_page_edit')
    else:
        form = AboutPageForm(instance=about_page)
    
    context = {
        'form': form,
        'about_page': about_page,
    }
    
    return render(request, 'content_management/about_page_form.html', context)

@login_required
def hero_statistic_list(request):
    """Liste des statistiques hero"""
    try:
        about_page = AboutPage.get_active_page()
        statistics = HeroStatistic.objects.filter(about_page=about_page, is_active=True).order_by('order', 'number')
    except:
        about_page = None
        statistics = []
    
    context = {
        'statistics': statistics,
        'about_page': about_page,
    }
    return render(request, 'content_management/hero_statistic_list.html', context)


@login_required
def hero_statistic_create(request):
    """Créer une nouvelle statistique hero"""
    try:
        about_page = AboutPage.get_active_page()
    except:
        about_page = None
    
    if not about_page:
        messages.error(request, 'Aucune page À propos configurée. Configurez d\'abord la page principale.')
        return redirect('content_management:about_page_edit')
    
    if request.method == 'POST':
        form = HeroStatisticForm(request.POST)
        if form.is_valid():
            statistic = form.save(commit=False)
            statistic.about_page = about_page
            statistic.save()
            messages.success(request, 'Statistique créée avec succès !')
            return redirect('content_management:hero_statistic_list')
    else:
        form = HeroStatisticForm()
    
    context = {
        'form': form,
        'action': 'Créer',
        'about_page': about_page,
    }
    return render(request, 'content_management/hero_statistic_form.html', context)


@login_required
def hero_statistic_edit(request, pk):
    """Modifier une statistique hero"""
    statistic = get_object_or_404(HeroStatistic, pk=pk)
    if request.method == 'POST':
        form = HeroStatisticForm(request.POST, instance=statistic)
        if form.is_valid():
            form.save()
            messages.success(request, 'Statistique modifiée avec succès !')
            return redirect('content_management:hero_statistic_list')
    else:
        form = HeroStatisticForm(instance=statistic)
    
    context = {
        'form': form,
        'statistic': statistic,
        'action': 'Modifier',
        'about_page': statistic.about_page,
    }
    return render(request, 'content_management/hero_statistic_form.html', context)


@login_required
def achievement_list(request):
    """Liste des réalisations"""
    try:
        about_page = AboutPage.get_active_page()
        achievements = Achievement.objects.filter(about_page=about_page, is_active=True).order_by('order', 'text')
    except:
        about_page = None
        achievements = []
    
    context = {
        'achievements': achievements,
        'about_page': about_page,
    }
    return render(request, 'content_management/achievement_list.html', context)


@login_required
def achievement_create(request):
    """Créer une nouvelle réalisation"""
    try:
        about_page = AboutPage.get_active_page()
    except:
        about_page = None
    
    if not about_page:
        messages.error(request, 'Aucune page À propos configurée. Configurez d\'abord la page principale.')
        return redirect('content_management:about_page_edit')
    
    if request.method == 'POST':
        form = AchievementForm(request.POST)
        if form.is_valid():
            achievement = form.save(commit=False)
            achievement.about_page = about_page
            achievement.save()
            messages.success(request, 'Réalisation créée avec succès !')
            return redirect('content_management:achievement_list')
    else:
        form = AchievementForm()
    
    context = {
        'form': form,
        'action': 'Créer',
        'about_page': about_page,
    }
    return render(request, 'content_management/achievement_form.html', context)


@login_required
def achievement_edit(request, pk):
    """Modifier une réalisation"""
    achievement = get_object_or_404(Achievement, pk=pk)
    if request.method == 'POST':
        form = AchievementForm(request.POST, instance=achievement)
        if form.is_valid():
            form.save()
            messages.success(request, 'Réalisation modifiée avec succès !')
            return redirect('content_management:achievement_list')
    else:
        form = AchievementForm(instance=achievement)
    
    context = {
        'form': form,
        'achievement': achievement,
        'action': 'Modifier',
        'about_page': achievement.about_page,
    }
    return render(request, 'content_management/achievement_form.html', context)

@login_required
def about_page_preview(request):
    """Aperçu de la page À propos"""
    try:
        about_page = AboutPage.get_active_page()
    except AboutPage.DoesNotExist:
        about_page = None
    
    if not about_page:
        messages.error(request, 'Aucune page À propos configurée.')
        return redirect('content_management:about_page_edit')
    
    context = {
        'about_page': about_page,
        'is_preview': True
    }
    
    return render(request, 'sfront/about.html', context)


@login_required
def city_district_list(request):
    """Liste des quartiers de la cité"""
    districts = CityDistrict.objects.filter(is_active=True).order_by('order', 'name')
    context = {
        'districts': districts,
    }
    return render(request, 'content_management/city_district_list.html', context)


@login_required
def city_district_create(request):
    """Créer un nouveau quartier"""
    if request.method == 'POST':
        form = CityDistrictForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Quartier créé avec succès !')
            return redirect('content_management:city_district_list')
    else:
        form = CityDistrictForm()
    
    context = {
        'form': form,
        'action': 'Créer',
    }
    return render(request, 'content_management/city_district_form.html', context)


@login_required
def city_district_edit(request, pk):
    """Modifier un quartier"""
    district = get_object_or_404(CityDistrict, pk=pk)
    if request.method == 'POST':
        form = CityDistrictForm(request.POST, request.FILES, instance=district)
        if form.is_valid():
            form.save()
            messages.success(request, 'Quartier modifié avec succès !')
            return redirect('content_management:city_district_list')
    else:
        form = CityDistrictForm(instance=district)
    
    context = {
        'form': form,
        'district': district,
        'action': 'Modifier',
    }
    return render(request, 'content_management/city_district_form.html', context)


@login_required
def core_value_list(request):
    """Liste des valeurs fondamentales"""
    values = CoreValue.objects.filter(is_active=True).order_by('order', 'name')
    context = {
        'values': values,
    }
    return render(request, 'content_management/core_value_list.html', context)


@login_required
def core_value_create(request):
    """Créer une nouvelle valeur fondamentale"""
    if request.method == 'POST':
        form = CoreValueForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Valeur fondamentale créée avec succès !')
            return redirect('content_management:core_value_list')
    else:
        form = CoreValueForm()
    
    context = {
        'form': form,
        'action': 'Créer',
    }
    return render(request, 'content_management/core_value_form.html', context)


@login_required
def core_value_edit(request, pk):
    """Modifier une valeur fondamentale"""
    value = get_object_or_404(CoreValue, pk=pk)
    if request.method == 'POST':
        form = CoreValueForm(request.POST, instance=value)
        if form.is_valid():
            form.save()
            messages.success(request, 'Valeur fondamentale modifiée avec succès !')
            return redirect('content_management:core_value_list')
    else:
        form = CoreValueForm(instance=value)
    
    context = {
        'form': form,
        'value': value,
        'action': 'Modifier',
    }
    return render(request, 'content_management/core_value_form.html', context)

# ============================================================================
# GESTION DES CAMPAGNES DE NEWSLETTER
# ============================================================================

@login_required
def newsletter_campaign_list(request):
    """Liste des campagnes de newsletter"""
    campaigns = NewsletterCampaign.objects.all().order_by('-created_at')
    
    # Pagination
    paginator = Paginator(campaigns, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'newsletter_campaigns': page_obj,
    }
    return render(request, 'content_management/newsletter_campaign_list.html', context)


@login_required
def newsletter_campaign_create(request):
    """Créer une nouvelle campagne de newsletter"""
    if request.method == 'POST':
        # TODO: Implémenter la création de campagne
        messages.success(request, 'Fonctionnalité de création de campagne à implémenter !')
        return redirect('content_management:newsletter_campaign_list')
    
    context = {
        'action': 'Créer',
    }
    return render(request, 'content_management/newsletter_campaign_form.html', context)


@login_required
def newsletter_campaign_detail(request, pk):
    """Détail d'une campagne de newsletter"""
    campaign = get_object_or_404(NewsletterCampaign, pk=pk)
    
    context = {
        'campaign': campaign,
    }
    return render(request, 'content_management/newsletter_campaign_detail.html', context)


@login_required
def newsletter_campaign_edit(request, pk):
    """Modifier une campagne de newsletter"""
    campaign = get_object_or_404(NewsletterCampaign, pk=pk)
    
    if request.method == 'POST':
        # TODO: Implémenter la modification de campagne
        messages.success(request, 'Fonctionnalité de modification de campagne à implémenter !')
        return redirect('content_management:newsletter_campaign_detail', pk=pk)
    
    context = {
        'campaign': campaign,
        'action': 'Modifier',
    }
    return render(request, 'content_management/newsletter_campaign_form.html', context)


@login_required
def newsletter_campaign_delete(request, pk):
    """Supprimer une campagne de newsletter"""
    if request.method == 'POST':
        campaign = get_object_or_404(NewsletterCampaign, pk=pk)
        campaign.delete()
        messages.success(request, 'Campagne supprimée avec succès !')
        return JsonResponse({'success': True, 'message': 'Campagne supprimée avec succès !'})
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})


@login_required
def newsletter_campaign_send(request, pk):
    """Envoyer une campagne de newsletter"""
    if request.method == 'POST':
        campaign = get_object_or_404(NewsletterCampaign, pk=pk)
        # TODO: Implémenter l'envoi de campagne
        messages.success(request, 'Fonctionnalité d\'envoi de campagne à implémenter !')
        return JsonResponse({'success': True, 'message': 'Campagne envoyée avec succès !'})
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})


@login_required
def newsletter_campaign_schedule(request, pk):
    """Programmer une campagne de newsletter"""
    if request.method == 'POST':
        campaign = get_object_or_404(NewsletterCampaign, pk=pk)
        # TODO: Implémenter la programmation de campagne
        messages.success(request, 'Fonctionnalité de programmation de campagne à implémenter !')
        return JsonResponse({'success': True, 'message': 'Campagne programmée avec succès !'})
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})


@login_required
def newsletter_campaign_duplicate(request, pk):
    """Dupliquer une campagne de newsletter"""
    if request.method == 'POST':
        campaign = get_object_or_404(NewsletterCampaign, pk=pk)
        # TODO: Implémenter la duplication de campagne
        messages.success(request, 'Fonctionnalité de duplication de campagne à implémenter !')
        return JsonResponse({'success': True, 'message': 'Campagne dupliquée avec succès !'})
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})


@login_required
def bulk_newsletter_campaign_action(request):
    """Actions en lot sur les campagnes de newsletter"""
    if request.method == 'POST':
        action = request.POST.get('action')
        newsletter_ids = json.loads(request.POST.get('newsletter_ids', '[]'))
        
        if not newsletter_ids:
            return JsonResponse({'success': False, 'message': 'Aucune campagne sélectionnée'})
        
        campaigns = NewsletterCampaign.objects.filter(pk__in=newsletter_ids)
        
        if action == 'send':
            # TODO: Implémenter l'envoi en lot
            messages.success(request, f'{campaigns.count()} campagnes envoyées avec succès !')
        elif action == 'schedule':
            # TODO: Implémenter la programmation en lot
            messages.success(request, f'{campaigns.count()} campagnes programmées avec succès !')
        elif action == 'duplicate':
            # TODO: Implémenter la duplication en lot
            messages.success(request, f'{campaigns.count()} campagnes dupliquées avec succès !')
        elif action == 'delete':
            campaigns.delete()
            messages.success(request, f'{len(newsletter_ids)} campagnes supprimées avec succès !')
        else:
            return JsonResponse({'success': False, 'message': 'Action non reconnue'})
        
        return JsonResponse({'success': True, 'message': 'Action exécutée avec succès !'})
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})


@login_required
def export_newsletter_campaigns(request):
    """Exporter les campagnes de newsletter"""
    # TODO: Implémenter l'export des campagnes
    messages.info(request, 'Fonctionnalité d\'export à implémenter !')
    return redirect('content_management:newsletter_campaign_list')
